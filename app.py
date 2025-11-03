import os
import tempfile
import pickle
from datetime import datetime
import pandas as pd
import yaml
from tabulate import tabulate
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, session, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

# ÁßªÈô§Â§öË™ûÁ≥ªÊîØÊåÅ
# from translations import get_text, get_language_options
from analyzers.clients.gemini_client import call_gemini as call_gemini_api
from langdetect import detect, DetectorFactory
import pdfplumber
import docx
from PIL import Image
import pytesseract
from bs4 import BeautifulSoup
import email
import email.policy
import re, json
import concurrent.futures
import threading
import uuid
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
import logging
import sqlite3
from typing import List, Dict, Any, Optional
# Configure Poppler path for pdf2image on Windows
# Configure Poppler path for Windows
import platform

if platform.system() == 'Windows':
    poppler_path = r'C:\Program Files\poppler-24.08.0\Library\bin'
    if os.path.exists(poppler_path):
        os.environ['PATH'] += os.pathsep + poppler_path

    # Configure Tesseract path for OCR
    tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    if os.path.exists(tesseract_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path
# Ë®≠ÁΩÆÊó•Ë™åÈÖçÁΩÆ
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

DetectorFactory.seed = 0

app = Flask(__name__)
app.secret_key = 'fileanalyzer_secret_key'

# ÁßªÈô§Â§öË™ûÁ≥ªÊ®°ÊùøÈÅéÊøæÂô®
# @app.template_filter('t')
# def translate_filter(key, **kwargs):
#     """Ê®°Êùø‰∏≠ÁöÑÁøªË≠ØÈÅéÊøæÂô®"""
#     lang = session.get('language', 'zh')
#     return get_text(key, lang, **kwargs)

# @app.context_processor
# def inject_language():
#     """Ê≥®ÂÖ•Ë™ûË®ÄÁõ∏ÈóúËÆäÊï∏Âà∞ÊâÄÊúâÊ®°Êùø"""
#     return {
#         'current_language': session.get('language', 'zh'),
#         'language_options': get_language_options()
#     }

# Set upload folder
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, mode=0o755)
    app.logger.info(f"Created upload folder: {UPLOAD_FOLDER}")
# Á¢∫‰øù‰∏äÂÇ≥ÁõÆÈåÑÊúâÊ≠£Á¢∫Ê¨äÈôê
try:
    os.chmod(UPLOAD_FOLDER, 0o755)
    app.logger.info(f"Set permissions for upload folder: {UPLOAD_FOLDER}")
except Exception as e:
    app.logger.warning(f"Failed to set permissions for upload folder: {e}")

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB ÈôêÂà∂

HISTORY_FILE = 'history.pkl'

# ËÆÄÂèñ config.yaml
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

# Â∞éÂÖ• prompts Âæû analyzers/services/prompts
try:
    from analyzers.services.prompts import SERVICE_PROMPTS
    # Â∞á SERVICE_PROMPTS Âêà‰ΩµÂà∞ config ‰∏≠
    config['prompts'] = SERVICE_PROMPTS
except ImportError as e:
    print(f"Warning: Could not import SERVICE_PROMPTS: {e}")
    # Â¶ÇÊûúÂ∞éÂÖ•Â§±ÊïóÔºå‰øùÊåÅÂéüÊúâÁöÑ config['prompts']

# ÁßªÈô§Â§öË™ûÁ≥ªÊîØÊåÅÔºåÂè™‰øùÁïôËã±Êñá
# LANGS = list(config['langs'].keys())

# History records
if os.path.exists(HISTORY_FILE):
    with open(HISTORY_FILE, 'rb') as f:
        history = pickle.load(f)
else:
    history = []

# Ëã•ÊúâÂÆâË£ù doclingÔºåÂèØÁî®ÊñºÊõ¥Âº∑Â§ßÁöÑÊñá‰ª∂Ëß£Êûê
try:
    from docling.document_converter import DocumentConverter
    DOC_CONVERTER = DocumentConverter()
except ImportError:
    DOC_CONVERTER = None

# Note: Moved to after init_db definition to avoid static analyzer false positive order issues

# Progress tracking
progress_store = {}
progress_lock = threading.Lock()

# Global variable for storing VESSEL and VOY.NO. information
vessel_voy_data = None

def update_progress(task_id, current, total, message=None):
    """Êõ¥Êñ∞‰ªªÂãôÈÄ≤Â∫¶"""
    try:
        percentage = (current / total * 100) if total > 0 else 0
        with progress_lock:
            progress_store[task_id] = {
                'current': current,
                'total': total,
                'percentage': percentage,
                'message': message or f'Completed {current}/{total} files'
            }
    except Exception as e:
        app.logger.error(f"Error updating progress for task {task_id}: {str(e)}")

def get_progress(task_id):
    """ÂèñÂæó‰ªªÂãôÈÄ≤Â∫¶"""
    try:
        with progress_lock:
            progress = progress_store.get(task_id, {
                'current': 0,
                'total': 1,
                'percentage': 0,
                'message': 'Task not found'
            })
            
            # Ë®òÈåÑÈÄ≤Â∫¶Ë´ãÊ±Ç
            app.logger.info(f"Progress request for task {task_id}: has_final_result={'final_result' in progress}")
            
            # Â¶ÇÊûú‰ªªÂãôÂÆåÊàê‰∏îÊúâÊúÄÁµÇÁµêÊûúÔºåËøîÂõûÊúÄÁµÇÁµêÊûú
            if 'final_result' in progress:
                app.logger.info(f"Returning final result for task {task_id}: {progress['final_result'].get('success', 'unknown')}")
                return jsonify(progress['final_result'])
            
            return jsonify(progress)
    except Exception as e:
        app.logger.error(f"Error getting progress for task {task_id}: {str(e)}")
        return jsonify({
            'current': 0,
            'total': 1,
            'percentage': 0,
            'message': f'Error retrieving progress: {str(e)}'
        }), 500

# =========================
# User and Permission Management: SQLite Simple Management
# =========================
DB_PATH = os.path.join(os.path.dirname(__file__), 'app.db')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_user_by_username(username):
    """Get user information by username"""
    conn = get_db_connection()
    try:
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if user:
            return dict(user)
        return None
    except Exception as e:
        logging.error(f"Failed to get user information: {e}")
        return None
    finally:
        conn.close()

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    # users Âü∫Êú¨Ë≥áË®ä
    cur.execute(
        '''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_admin INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1
        )'''
    )
    # User permissions for analyzers and VLM settings
    cur.execute(
        '''CREATE TABLE IF NOT EXISTS user_analyzers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            analyzer TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 1,
            vlm_provider TEXT NOT NULL DEFAULT 'cloud', -- 'cloud' or 'local'
            ollama_model TEXT DEFAULT NULL,
            ocr_lang TEXT DEFAULT 'auto',
            save_files INTEGER NOT NULL DEFAULT 0,
            UNIQUE(user_id, analyzer),
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        )'''
    )
    
    # Add save_files column if it doesn't exist (for existing databases)
    try:
        cur.execute('ALTER TABLE user_analyzers ADD COLUMN save_files INTEGER NOT NULL DEFAULT 0')
    except:
        pass  # Column already exists
    conn.commit()
    # Default admin account
    cur.execute("SELECT id FROM users WHERE username=?", ('admin',))
    row = cur.fetchone()
    if not row:
        cur.execute(
            "INSERT INTO users (username, password_hash, is_admin, is_active) VALUES (?, ?, ?, ?)",
            ('admin', generate_password_hash('admin123', method='pbkdf2:sha256'), 1, 1)
        )
        conn.commit()
    conn.close()
    
# Initialize database at startup (init_db already defined here)
try:
    init_db()
except Exception as e:
    logging.error(f"Failed to initialize database: {e}")

def get_current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    conn.close()
    return row

def calculate_user_statistics(history_records):
    """Calculate user statistics including total tokens and daily usage"""
    from collections import defaultdict
    from datetime import datetime
    import re
    
    user_stats = defaultdict(lambda: {
        'total_tokens': 0,
        'daily_usage': defaultdict(int),
        'total_records': 0,
        'total_files': 0,
        'total_processed': 0
    })
    
    for record in history_records:
        username = record.get('username', 'Unknown User')
        tokens = record.get('tokens', 0)
        
        if isinstance(tokens, (int, float)):
            tokens = int(tokens)
        else:
            tokens = 0
            
        user_stats[username]['total_tokens'] += tokens
        user_stats[username]['total_records'] += 1
        
        # Count files
        files = record.get('files', [])
        if files:
            user_stats[username]['total_files'] += len(files)
        
        # Count processing results (extract row count from result field)
        result_str = record.get('result', '')
        if result_str:
            # Try to extract row count from result, e.g. "24 rows 13 columns" -> 24
            match = re.search(r'(\d+)Ë°å', result_str)
            if match:
                rows = int(match.group(1))
                user_stats[username]['total_processed'] += rows
            else:
                # If no row count info, calculate by file count
                user_stats[username]['total_processed'] += len(files) if files else 1
        
        # Parse date
        time_str = record.get('time') or record.get('timestamp', '')
        if time_str:
            try:
                # Try to parse date part (YYYY-MM-DD)
                if ' ' in time_str:
                    date_part = time_str.split(' ')[0]
                else:
                    date_part = time_str
                
                # Validate date format
                datetime.strptime(date_part, '%Y-%m-%d')
                user_stats[username]['daily_usage'][date_part] += tokens
            except ValueError:
                # Skip if date format is invalid
                pass
    
    return dict(user_stats)

def create_bilingual_log(log_entries_zh, log_filename):
    """Create English-Chinese bilingual log file (English version first, Chinese version second)"""
    import re
    
    # English-Chinese translation mapping
    translations = {
        '=== Ê™îÊ°àÂàÜÊûêËôïÁêÜÊó•Ë™å ===': '=== File Analysis Processing Log ===',
        'ÈñãÂßãÊôÇÈñì': 'Start Time',
        'ÂàÜÊûêÈ°ûÂûã': 'Analysis Type',
        'Ë™ûË®Ä': 'Language',
        'Ê™îÊ°àÊï∏Èáè': 'File Count',
        'Ëá™ÂÆöÁæ©ÊèêÁ§∫Ë©û': 'Custom Prompt',
        'ÊòØ': 'Yes',
        'Âê¶': 'No',
        '=== OCR ÊñáÂ≠óÊèêÂèñÈöéÊÆµ ===': '=== OCR Text Extraction Phase ===',
        'Ê™îÊ°àÊ∫ñÂÇô': 'File Preparation',
        'OCR ÊèêÂèñ': 'OCR Extraction',
        'È†ÅÊï∏': 'Pages',
        'ÊèêÂèñÊñáÂ≠óÈï∑Â∫¶': 'Extracted Text Length',
        'Â≠óÁ¨¶': 'characters',
        'ÁãÄÊÖã': 'Status',
        'ÊàêÂäü': 'Success',
        'Â§±Êïó': 'Failed',
        '=== LLM ÂàÜÊûêÈöéÊÆµ ===': '=== LLM Analysis Phase ===',
        'LLM ÂàÜÊûê': 'LLM Analysis',
        'Tokens ‰ΩøÁî®': 'Tokens Used',
        'ÂàÜÊûêÁµêÊûú': 'Analysis Result',
        'VLM Ë£úÊïëÊ¨Ñ‰Ωç': 'VLM Remediated Fields',
        'VLM Tokens ‰ΩøÁî®': 'VLM Tokens Used',
        'VLM Ë£úÊïë': 'VLM Remediation',
        'ÁÑ°ÈúÄË¶Å': 'Not Required',
        '=== ËôïÁêÜÊëòË¶Å ===': '=== Processing Summary ===',
        'Á∏ΩËôïÁêÜÊôÇÈñì': 'Total Processing Time',
        'Áßí': 'seconds',
        'Á∏Ω Tokens ‰ΩøÁî®': 'Total Tokens Used',
        'Á∏ΩÈ†ÅÊï∏': 'Total Pages',
        'ÊàêÂäüËôïÁêÜÊ™îÊ°à': 'Successfully Processed Files',
        '=== Ë©≥Á¥∞Ê™îÊ°àËôïÁêÜÁµ±Ë®à ===': '=== Detailed File Processing Statistics ===',
        'Ê™îÊ°à': 'File',
        'LLM Tokens': 'LLM Tokens',
        'VLM Tokens': 'VLM Tokens',
        'Á∏Ω Tokens': 'Total Tokens',
        'ÊàêÂäüÊèêÂèñÊ¨Ñ‰Ωç': 'Successfully Extracted Fields',
        'Áº∫Â§±Ê¨Ñ‰Ωç': 'Missing Fields',
        'VLM Ë£úÊïëÊ¨Ñ‰Ωç': 'VLM Remediated Fields',
        'ÁµêÊùüÊôÇÈñì': 'End Time',
        'Ëº∏Âá∫ Excel': 'Output Excel',
        'Ëº∏Âá∫ Log': 'Output Log',
        'ÂÄã': '',
        ' bytes': ' bytes'
    }
    
    # Special pattern translations (for handling compound words)
    pattern_translations = [
        (r'ÊàêÂäüÊèêÂèñÊ¨Ñ‰Ωç', 'Successfully Extracted Fields'),
        (r'ÊàêÂäüËôïÁêÜÊ™îÊ°à', 'Successfully Processed Files'),
        (r'Á∏Ω([A-Za-z\s]+):', r'Total \1:'),
        (r'ÊàêÂäü([A-Za-z\s]+)', r'Successfully \1'),
        (r'Ê™îÊ°à([A-Za-z\s]*)', r'File\1'),
        (r'([A-Za-z\s]+)Ê¨Ñ‰Ωç', r'\1 Fields'),
        (r'([A-Za-z\s]+)È†ÅÊï∏', r'\1 Pages'),
        (r'([A-Za-z\s]+)ÂàÜÊûê', r'\1 Analysis')
    ]
    
    # Create English version of log entries
    log_entries_en = []
    for entry in log_entries_zh:
        en_entry = entry
        
        # First perform basic translation
        for zh_text, en_text in translations.items():
            if zh_text in en_entry:
                en_entry = en_entry.replace(zh_text, en_text)
        
        # Then perform pattern translation
        for pattern, replacement in pattern_translations:
            en_entry = re.sub(pattern, replacement, en_entry)
        
        # Handle special Chinese-English mixed cases
        # Remove remaining Chinese characters but keep numbers and symbols
        def clean_mixed_text(text):
            # Protect Chinese characters in filenames
            if '.pdf' in text or '.xlsx' in text:
                return text
            
            # Special handling for statistics rows
            if '  - ' in text or '    * ' in text:
                # Separate indentation, Chinese part and English/number part
                indent_match = re.match(r'^(\s*[-*]\s*)', text)
                if indent_match:
                    indent = indent_match.group(1)
                    content = text[len(indent):]
                    
                    # If content contains Chinese, try to clean further
                    if re.search(r'[\u4e00-\u9fff]', content):
                        # Keep numbers, English, colons, spaces, etc.
                        cleaned_content = re.sub(r'[\u4e00-\u9fff]+', '', content)
                        # Clean up extra spaces
                        cleaned_content = re.sub(r'\s+', ' ', cleaned_content).strip()
                        if cleaned_content:
                            return indent + cleaned_content
                    return text
            
            return text
        
        en_entry = clean_mixed_text(en_entry)
        log_entries_en.append(en_entry)
    
    # Build bilingual log: English version first, Chinese version second
    bilingual_log = []
    
    # English version
    bilingual_log.append("=" * 80)
    bilingual_log.append("ENGLISH VERSION")
    bilingual_log.append("=" * 80)
    bilingual_log.extend(log_entries_en)
    
    bilingual_log.append("")
    bilingual_log.append("")
    
    # Chinese version
    bilingual_log.append("=" * 80)
    bilingual_log.append("‰∏≠ÊñáÁâàÊú¨")
    bilingual_log.append("=" * 80)
    bilingual_log.extend(log_entries_zh)
    
    return bilingual_log

def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect(url_for('login', next=request.path))
        if user['is_active'] != 1:
            session.clear()
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user or user['is_admin'] != 1:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return wrapper

def get_user_analyzer_setting(user_id: int, analysis_type: str):
    conn = get_db_connection()
    row = conn.execute(
        'SELECT enabled, vlm_provider, ollama_model, ocr_lang, save_files FROM user_analyzers WHERE user_id=? AND analyzer=?',
        (user_id, analysis_type)
    ).fetchone()
    conn.close()
    if not row:
        # If not set, default to enabled and use cloud
        return {'enabled': True, 'vlm_provider': 'cloud', 'ollama_model': None, 'ocr_lang': 'auto', 'save_files': False}
    return {'enabled': bool(row['enabled']), 'vlm_provider': row['vlm_provider'], 'ollama_model': row['ollama_model'], 'ocr_lang': row['ocr_lang'] or 'auto', 'save_files': bool(row['save_files'])}

def get_user_allowed_analyzers(user_id: int):
    # Return list of analyzers enabled by user. If no settings, default to allow all prompts keys in config
    conn = get_db_connection()
    rows = conn.execute('SELECT analyzer FROM user_analyzers WHERE user_id=? AND enabled=1', (user_id,)).fetchall()
    conn.close()
    if not rows:
        return list(config.get('prompts', {}).keys())
    return [r['analyzer'] for r in rows]

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password) and user['is_active'] == 1:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = user['is_admin']
            next_url = request.args.get('next') or url_for('index')
            return redirect(next_url)
        return render_template('login.html', error='Invalid username or password, or account is disabled')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def save_history():
    with open(HISTORY_FILE, 'wb') as f:
        pickle.dump(history, f)

def extract_json_from_response(resp):
    resp = resp.strip()
    # Remove markdown markers
    if resp.startswith('```json'):
        resp = resp[7:]
    if resp.startswith('```'):
        resp = resp[3:]
    if resp.endswith('```'):
        resp = resp[:-3]
    resp = resp.strip()
    # Try direct parsing
    try:
        return json.loads(resp)
    except Exception:
        pass
    # Try to extract array
    match = re.search(r'\[.*\]', resp, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            pass
    # Try to extract first object
    match = re.search(r'\{[\s\S]*\}', resp)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            pass
    # If all fail, return original but provide clearer error message
    return {
        'error': 'Gemini API response is not in JSON format',
        'response': resp[:500] + '...' if len(resp) > 500 else resp
    }

def extract_text_from_pdf(file_path):
    """Extract text from PDF, use OCR if extraction fails"""
    try:
        texts = []
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text and text.strip():
                    texts.append(text)
                else:
                    images = convert_from_path(
                        file_path,
                        first_page=i + 1,
                        last_page=i + 1,
                        poppler_path=r'C:\Program Files\poppler-24.08.0\Library\bin'
                    )
                    ocr_lang = config.get('ocr_lang', 'eng') or 'eng'
                    for image in images:
                        ocr_text = pytesseract.image_to_string(image, lang=ocr_lang)
                        texts.append(ocr_text)
        result = '\n'.join(texts)
        app.logger.info(f"extract_text_from_pdf: {file_path} Final return length: {len(result) if result else 0}")
        return result
    except Exception as e:
        app.logger.error(f"PDF extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_word(file_path):
    """Extract text from Word document"""
    try:
        doc = docx.Document(file_path)
        texts = []
        for paragraph in doc.paragraphs:
            texts.append(paragraph.text)
        return '\n'.join(texts)
    except Exception as e:
        app.logger.error(f"Word extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_excel(file_path):
    """Extract text from Excel"""
    try:
        df = pd.read_excel(file_path)
        return df.to_string()
    except Exception as e:
        app.logger.error(f"Excel extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_csv(file_path):
    """Extract text from CSV"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        app.logger.error(f"CSV extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_image(file_path):
    """Extract text from image (OCR)"""
    try:
        image = Image.open(file_path)
        ocr_lang = config.get('ocr_lang', 'eng') or 'eng'
        text = pytesseract.image_to_string(image, lang=ocr_lang)
        return text
    except Exception as e:
        app.logger.error(f"Image OCR error {file_path}: {str(e)}")
        return None

def extract_file_content(file_path):
    """Extract file content"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.pdf':
            return extract_text_from_pdf(file_path)
        elif file_ext in ['.doc', '.docx']:
            return extract_text_from_word(file_path)
        elif file_ext in ['.xls', '.xlsx']:
            return extract_text_from_excel(file_path)
        elif file_ext == '.csv':
            return extract_text_from_csv(file_path)
        elif file_ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_ext in ['.jpg', '.jpeg', '.png', '.bmp']:
            return extract_text_from_image(file_path)
        else:
            app.logger.warning(f"Unsupported file type: {file_ext}")
            return None
            
    except Exception as e:
        app.logger.error(f"Error occurred while extracting file content {file_path}: {str(e)}")
        return None

from analyzers.services import analyze_text_with_prompt_with_gemini

def export_to_excel(data, filename, preview_title=None, manual_data=None, include_logo=False, user_info=None, vessel_voy_info=None, keep_filename=False):
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        
        # Êï∏ÊìöÂ∑≤Á∂ìÂú® cargo_bl_postprocess ‰∏≠ËôïÁêÜÈÅé‰∫ÜÔºåÁõ¥Êé•‰ΩøÁî®
        processed_data = data
        app.logger.info(f"export_to_excel: keep_filename={keep_filename}")
        app.logger.info(f"export_to_excel: data keys sample={[list(item.keys()) if isinstance(item, dict) else 'not_dict' for item in data[:2]]}")
        
        df = pd.DataFrame(processed_data)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        df.to_excel(excel_path, index=False)
        wb = load_workbook(excel_path)
        ws = wb.active
        current_row = 1
        
        # 1. Ê®ôÈ°åÂíå Logo ÂêåË°å
        if preview_title:
            today_str = datetime.now().strftime('%Y-%m-%d')
            
            # ÊèíÂÖ•Ê®ôÈ°åË°å
            ws.insert_rows(current_row, 1)
            
            # Ë®≠ÁΩÆÊ®ôÈ°åË°åÁöÑÊ∑°ËóçËâ≤Â∫ïËâ≤
            light_blue_fill = PatternFill('solid', fgColor='E6F3FF')
            
            # Ê®ôÈ°åÊñáÂ≠óÔºàÂ∑¶ÈÇäÂíå‰∏≠ÈñìÔºâ
            if ws.max_column > 1:
                # Âêà‰ΩµÂâçÂπæÊ¨ÑÁµ¶Ê®ôÈ°å
                merge_end_col = max(1, ws.max_column - 2) if include_logo else ws.max_column
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=merge_end_col)
            
            title_cell = ws.cell(row=current_row, column=1)
            # ‰ΩøÁî®Áî®Êà∂ÂêçÁ®±ÔºåÂ¶ÇÊûúÊ≤íÊúâÂâá‰ΩøÁî®ÈªòË™çÊ®ôÈ°å
            if user_info and user_info.get('display_name'):
                title_cell.value = user_info['display_name']
            else:
                title_cell.value = preview_title
            title_cell.font = Font(size=20, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = light_blue_fill
            
            # LogoÔºàÂè≥ÈÇäÔºâ
            if include_logo:
                # ÂÑ™ÂÖà‰ΩøÁî®Áî®Êà∂ÁöÑlogoÔºåÂ¶ÇÊûúÊ≤íÊúâÂâá‰ΩøÁî®ÈªòË™çlogo
                logo_path = None
                if user_info and user_info.get('logo_file'):
                    user_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], user_info['logo_file'])
                    if os.path.exists(user_logo_path):
                        logo_path = user_logo_path
                
                # Â¶ÇÊûúÊ≤íÊúâÁî®Êà∂logoÔºå‰ΩøÁî®ÈªòË™çlogo
                if not logo_path:
                    default_logo_path = os.path.join('static', 'hansa.png')
                    if os.path.exists(default_logo_path):
                        logo_path = default_logo_path
                
                if logo_path:
                    img = OpenpyxlImage(logo_path)
                    img.width = 60
                    img.height = 60
                    # Ë®≠ÁΩÆË°åÈ´ò‰ª•ÂÆπÁ¥ç logo
                    ws.row_dimensions[current_row].height = 50
                    # Â∞á logo ÊîæÂú®ÊúÄÂæå‰∏ÄÊ¨Ñ
                    last_col_letter = get_column_letter(ws.max_column)
                    ws.add_image(img, f'{last_col_letter}{current_row}')
                    
                    # ÁÇ∫ logo ÊâÄÂú®ÁöÑÊ¨Ñ‰Ωç‰πüË®≠ÁΩÆÂ∫ïËâ≤
                    logo_cell = ws.cell(row=current_row, column=ws.max_column)
                    logo_cell.fill = light_blue_fill
            
            # ÁÇ∫Êï¥Ë°åË®≠ÁΩÆÂ∫ïËâ≤
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=current_row, column=col)
                if cell.fill.fgColor.rgb != light_blue_fill.fgColor.rgb:
                    cell.fill = light_blue_fill
            
            current_row += 1
            
            # Êó•ÊúüË°å
            ws.insert_rows(current_row, 1)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
            date_cell = ws.cell(row=current_row, column=1)
            date_cell.value = today_str
            date_cell.font = Font(size=12, italic=True)
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # VESSEL Âíå VOY.NO. Ë°åÔºàÈáùÂ∞ç Cargo_BLÔºâ
            if vessel_voy_info:
                # VESSEL Ë°å
                ws.insert_rows(current_row, 1)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
                vessel_cell = ws.cell(row=current_row, column=1)
                vessel_cell.value = f"VESSEL: {vessel_voy_info.get('VESSEL', '')}"
                vessel_cell.font = Font(size=12, bold=True)
                vessel_cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
                
                # VOY.NO. Ë°å
                ws.insert_rows(current_row, 1)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
                voy_cell = ws.cell(row=current_row, column=1)
                voy_cell.value = f"VOY.NO.: {vessel_voy_info.get('VOY.NO.', '')}"
                voy_cell.font = Font(size=12, bold=True)
                voy_cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
        
        # 2. ÊâãÂãïÊèíÂÖ•ÁöÑË≥áÊñôÔºàËôïÁêÜÁµêÊßãÂåñÊï∏ÊìöÔºâ
        if manual_data and manual_data.get('rows'):
            manual_rows = manual_data['rows']
            if manual_rows:
                rows_to_insert = len(manual_rows)
                ws.insert_rows(current_row, rows_to_insert)
                
                for row_idx, row_data in enumerate(manual_rows):
                    for col_idx, cell_data in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row + row_idx, column=col_idx)
                        
                        # ËôïÁêÜÁµêÊßãÂåñÁöÑ cell_data
                        if isinstance(cell_data, dict):
                            cell.value = cell_data.get('value', '')
                            color = cell_data.get('color', '#000000')
                            font_size = cell_data.get('fontSize', 12)
                        else:
                            cell.value = cell_data
                            color = '#000000'
                            font_size = 12
                        
                        # Ë®≠ÁΩÆÂ≠óÈ´îÊ®£Âºè
                        cell.font = Font(bold=True, size=font_size, color=color.replace('#', ''))
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += rows_to_insert
        
        # 3. ‰∏ªË¶ÅË≥áÊñôÁöÑÊ®ôÈ°åË°åÊ®£Âºè
        header_row = current_row
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='000000')
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 4. Ê¨ÑÂØ¨Ëá™ÂãïË™øÊï¥ÔºàÊ†πÊìöÂÖßÂÆπÂíåÂ≠óÈ´îÂ§ßÂ∞èÔºâ
        for col, cell in enumerate(ws[header_row], 1):
            # Ë®àÁÆóË©≤Ê¨ÑÁöÑÊúÄÂ§ßÂÖßÂÆπÈï∑Â∫¶
            max_len = max(len(str(cell.value)),
                          max((len(str(ws.cell(row=row, column=col).value)) for row in range(header_row+1, ws.max_row+1)), default=0))
            
            # Ë®àÁÆóË©≤Ê¨ÑÁöÑÊúÄÂ§ßÂ≠óÈ´îÂ§ßÂ∞èÔºåËôïÁêÜ None ÂÄº
            font_sizes = []
            for row in range(1, ws.max_row+1):
                font_size = getattr(ws.cell(row=row, column=col).font, 'size', 12)
                if font_size is not None:
                    font_sizes.append(font_size)
                else:
                    font_sizes.append(12)
            max_font = max(font_sizes, default=12)
            
            # Ê†πÊìöÂ≠óÈ´îÂ§ßÂ∞èË™øÊï¥Ê¨ÑÂØ¨
            adjusted_width = max(10, min(max_len * max_font / 12 + 2, 60))
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width
        
        # 5. Total row Âíå Cgos row Ê®£Âºè
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            first_cell = row[0].value
            if str(first_cell).strip().lower() == 'total':
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='E5C29F')
            if str(first_cell).strip().startswith('# Cgos'):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='C6EFCE')
        
        wb.save(excel_path)
        return excel_path
    except Exception as e:
        app.logger.error(f"ÂåØÂá∫ Excel ÈåØË™§: {str(e)}")
        return None

def save_to_history(entry):
    try:
        history_file = 'history.pkl'
        global history
        # ËºâÂÖ•ÁèæÊúâÊ≠∑Âè≤Ë®òÈåÑ
        if os.path.exists(history_file):
            with open(history_file, 'rb') as f:
                file_history = pickle.load(f)
        else:
            file_history = []
        # Ê∑ªÂä†Êñ∞Ë®òÈåÑ
        file_history.append(entry)
        # ÂÑ≤Â≠òÊ≠∑Âè≤Ë®òÈåÑ
        with open(history_file, 'wb') as f:
            pickle.dump(file_history, f)
        # ÂêåÊ≠•Âà∞ÂÖ®Âüü history
        history.append(entry)
    except Exception as e:
        app.logger.error(f"Failed to save history record: {str(e)}")

# ‰øÆÊ≠£Ê™îÊ°àÂÖßÂÆπÊèêÂèñÂáΩÊï∏ÂêçÁ®±
def extract_text(file_path, file_extension):
    """ÊèêÂèñÊ™îÊ°àÊñáÂ≠óÂÖßÂÆπ"""
    return extract_file_content(file_path)

@app.route('/', methods=['GET'])
@login_required
def index():
    lang = 'en'  # Default to English
    user = get_current_user()
    allowed = get_user_allowed_analyzers(user['id']) if user else list(config.get('prompts', {}).keys())
    # Filter config['prompts'] to show only user-available analyzers
    filtered_config = dict(config)
    filtered_config['prompts'] = {k: v for k, v in config.get('prompts', {}).items() if k in allowed}
    first_prompt_key = next(iter(filtered_config['prompts'].keys())) if filtered_config['prompts'] else ''
    return render_template('index.html', config=filtered_config, lang=lang, 
                         first_prompt_key=first_prompt_key, allowed_analyzers=allowed)

@app.route('/analyze', methods=['POST'])
@login_required
def analyze():
    try:
        # ÁîüÊàêÂîØ‰∏Ä‰ªªÂãôID
        task_id = str(uuid.uuid4())
        
        files = request.files.getlist('files')
        analysis_type = request.form.get('analysis_type')
        # 1Ô∏è‚É£ Get custom prompt from form
        custom_prompt = request.form.get('custom_prompt', '').strip()

        # 2Ô∏è‚É£ Fallback to main prompt if custom_prompt is empty
        if not custom_prompt:
            print("‚ö†Ô∏è custom_prompt empty, using main PROMPTS fallback")
            # SERVICE_PROMPTS['Cargo_BL'] contains the default prompts per language
            svc_prompts_code = SERVICE_PROMPTS.get('Cargo_BL', {})
            # Pick English ('en') or any other language you want
            custom_prompt = svc_prompts_code.get('en') or ""

        print(f"üß† ARJ Using custom_prompt (first 10 words): {' '.join(custom_prompt.split()[:8])}")

        keep_filename = request.form.get('keep_filename') == 'on'
        user = get_current_user()
        # Ê™¢Êü•‰ΩøÁî®ËÄÖÊòØÂê¶ÊúâË©≤ÂàÜÊûêÂô®Ê¨äÈôê
        allowed = get_user_allowed_analyzers(user['id'])
        if analysis_type not in allowed:
            return jsonify({'success': False, 'error': 'No permission to use this analyzer'}), 403

        # Ëß£ÊûêÁî®Êà∂Ë®≠ÂÆöÔºàÈÅøÂÖçÂú®ËÉåÊôØÂü∑Ë°åÁ∑í‰∏≠ÂèñÁî® sessionÔºâ
        setting = get_user_analyzer_setting(user['id'], analysis_type)
        vlm_provider = setting.get('vlm_provider', 'cloud')
        vlm_model = setting.get('ollama_model')
        lang = setting.get('ocr_lang', 'auto')  # ÂæûÁî®Êà∂Ë®≠ÂÆöÁç≤ÂèñË™ûË®Ä
        save_files = setting.get('save_files', False)  # ÂæûÁî®Êà∂Ë®≠ÂÆöÁç≤ÂèñÊòØÂê¶‰øùÂ≠òÊñá‰ª∂
        username = session.get('username')
        
        if not files or all(f.filename == '' for f in files):
            return jsonify({'success': False, 'error': 'Please select files'})
        
        # È†êËôïÁêÜÊ™îÊ°à - ÂÖàËÆÄÂèñÊâÄÊúâÊ™îÊ°àÂÖßÂÆπÈÅøÂÖç‰∏¶Ë°åËôïÁêÜ‰∏≠ÁöÑÊ™îÊ°àÈóúÈñâÂïèÈ°å
        file_data_list = []
        for file in files:
            if file.filename != '':
                try:
                    file.seek(0)
                    file_content = file.read()
                    file_data_list.append({
                        'filename': file.filename,
                        'content': file_content
                    })
                except Exception as e:
                    app.logger.error(f"Failed to read file {file.filename}: {str(e)}")
                    continue
        
        if not file_data_list:
            return jsonify({'success': False, 'error': 'Unable to read any files'})
        
        # Initialize progress
        total_files = len(file_data_list)
        update_progress(task_id, 0, total_files, 'Starting analysis...')
        
        # Run analysis in background
        thread = threading.Thread(
            target=process_files_with_progress, 
            args=(task_id, file_data_list, analysis_type, lang, custom_prompt, keep_filename, vlm_provider, vlm_model, username, save_files)
        )
        thread.start()
        
        return jsonify({'success': True, 'task_id': task_id})
    
    except Exception as e:
        app.logger.error(f"Analysis error: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

def extract_worker(args):
    filename, file_path = args
    import pdfplumber
    local_total_pages = 0
    content = None
    # Set Poppler path for Windows
    POPPLER_PATH = r"C:\Program Files\poppler-24.08.0\Library\bin"
    # Configure Tesseract path
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    print(f"[Worker] Processing file: {filename}")
    print(f"[Worker] File path: {file_path}")

    if filename.lower().endswith('.pdf'):
        try:
            with pdfplumber.open(file_path) as pdf:
                page_count = len(pdf.pages)
                local_total_pages = page_count
                # app.logger.info(f"PDF {filename} È†ÅÊï∏: {page_count}")  # Â§öÈÄ≤Á®ã‰∏çËÉΩÁî® app.logger
        except Exception as e:
            pass  # Â§öÈÄ≤Á®ã‰∏çËÉΩÁî® app.logger
    elif filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
        local_total_pages = 1
    from app import extract_file_content  # ‰øùË≠âÂ§öÈÄ≤Á®ãÂèØ import
    content = extract_file_content(file_path)
    return (filename, content, file_path, local_total_pages)


def process_files_with_progress(task_id, file_data_list, analysis_type, lang, custom_prompt, keep_filename=True,
                                vlm_provider: Optional[str] = None, vlm_model: Optional[str] = None,
                                username: Optional[str] = None, save_files: bool = False):
    import pdfplumber
    import time
    app.logger.warning(f"ARJ in the start of process_files_with_progress")
    try:
        saved_files = []
        start_time = time.time()
        total_files = len(file_data_list)
        processed_files = 0
        all_data = []
        total_pages = 0
        total_tokens = 0
        log_entries = []
        log_entries.append(f"=== File Analysis Processing Log ===")
        log_entries.append(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        log_entries.append(f"Analysis Type: {analysis_type}")
        log_entries.append(f"Language: {lang}")
        log_entries.append(f"File Count: {total_files}")
        log_entries.append(f"Custom Prompt: {'Yes' if custom_prompt else 'No'}")
        log_entries.append("")

        update_progress(task_id, 0, total_files, 'Starting analysis...')

        file_path_list = []
        for file_data in file_data_list:
            filename = file_data['filename']
            file_content = file_data['content']
            unique_filename = f"{uuid.uuid4().hex}_{secure_filename(filename)}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            with open(file_path, 'wb') as f:
                f.write(file_content)
                f.flush()
                os.fsync(f.fileno())
            try:
                os.chmod(file_path, 0o644)
            except Exception as e:
                app.logger.warning(f"Failed to set file permissions for {file_path}: {e}")
            file_path_list.append((filename, file_path))
            log_entries.append(f"File Preparation: {filename} ({len(file_content)} bytes)")

        log_entries.append("")
        log_entries.append("=== OCR Text Extraction Phase ===")

        extracted = []
        with ProcessPoolExecutor(max_workers=4) as executor:
            for idx, (filename, content, file_path, local_total_pages) in enumerate(
                    executor.map(extract_worker, file_path_list)):
                processed_files += 1
                update_progress(task_id, processed_files, total_files * 2, f"Analyzing {filename} ")
                extracted.append((filename, content, file_path))
                total_pages += local_total_pages
                content_length = len(content) if content else 0
                log_entries.append(f"OCR Extraction: {filename}")
                log_entries.append(f"  - Pages: {local_total_pages}")
                log_entries.append(f"  - Extracted Text Length: {content_length} characters")
                log_entries.append(f"  - Status: {'Success' if content_length > 0 else 'Failed'}")

        log_entries.append("")
        log_entries.append("=== LLM Analysis Phase ===")

        # In the file with process_files_with_progress function
        # Add this before the extract_worker section (around line 45)

        def extract_worker_multi_bl(file_path_tuple):
            """Modified extract worker for multi-BL PDFs"""
            filename, file_path = file_path_tuple

            # Check if PDF has multiple BLs
            if filename.lower().endswith('.pdf'):
                with pdfplumber.open(file_path) as pdf:
                    total_pages = len(pdf.pages)

                    # Check for multiple BLs markers
                    bl_markers = ['FIRST ORIGINAL', 'SECOND ORIGINAL', 'THIRD ORIGINAL']
                    bl_count = 0

                    full_text = ""
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        full_text += text + "\n"
                        for marker in bl_markers:
                            if marker in text:
                                bl_count += 1

                    # If multiple BLs detected, process differently
                    if bl_count > 4:  # More than one set of BLs
                        return process_multi_bl_separately(file_path, filename)
                    else:
                        # Single BL - use existing logic
                        return filename, full_text, file_path, total_pages

            # For non-PDF files, use existing logic
            return extract_worker(file_path_tuple)

        def process_multi_bl_separately(file_path, filename):
            """Process each BL in the PDF separately"""
            all_bl_texts = []

            with pdfplumber.open(file_path) as pdf:
                current_bl = ""
                bl_markers = ['FIRST ORIGINAL', 'SECOND ORIGINAL', 'THIRD ORIGINAL', 'COPY NOT NEGOTIABLE']

                for page in pdf.pages:
                    text = page.extract_text() or ""

                    # Check if new BL starts
                    is_new_bl = any(marker in text and 'Bill of Lading' in text for marker in bl_markers)

                    if is_new_bl and current_bl:
                        # Process previous BL
                        all_bl_texts.append(current_bl)
                        current_bl = text
                    else:
                        current_bl += "\n" + text

                # Add last BL
                if current_bl:
                    all_bl_texts.append(current_bl)
            app.logger.info(f"ARJ in process_multi_bl_separately")

            # Combine all BLs with clear separation
            combined_text = "\n\n===NEXT_BL===\n\n".join(all_bl_texts)
            return filename, combined_text, file_path, len(all_bl_texts)

        def create_llm_worker(keep_filename_param):
            def llm_worker(args):
                app.logger.info("ARJ in llm_worker")
                filename, content, file_path = args

                if not content:
                    print("‚ö†Ô∏è Alert: content is None or empty!")
                    return [], 0

                total_tokens = 0
                all_results = []

                # Split content into sections
                bl_sections = content.split("===NEXT_BL===") if "===NEXT_BL===" in content else [content]

                # Check if all sections are empty
                if not any(section.strip() for section in bl_sections):
                    print("‚ö†Ô∏è Alert: all sections are empty!")
                    return [], 0

                print(f"DEBUG ‚Üí Number of BL sections: {len(bl_sections)}")
                print("DEBUG ‚Üí Sections preview (first 200 chars each):")
                for idx, section in enumerate(bl_sections, 1):
                    print(f"Section {idx}: {section[:200]!r}")

                # Helper function to clean VOY.NO.
                # Helper function to clean VOY.NO.
                def clean_voy_no(voy_value, bl_number=None):
                    """Extract 4-digit voyage number, with fallback to BL number"""

                    # First try to clean the provided value
                    if voy_value and voy_value not in ['Not Available', 'null', None, '']:
                        # Check if value indicates "not found"
                        if not any(phrase in str(voy_value).lower() for phrase in
                                   ['not found', 'not available', 'n/a', 'Êâæ‰∏çÂà∞']):
                            # Extract digits
                            cleaned = ''.join(filter(str.isdigit, str(voy_value)))
                            # Must be exactly 4 digits
                            if len(cleaned) == 4:
                                return cleaned

                    # Fallback: Extract from BL number (YYMM format)
                    if bl_number:
                        # BL format: "EMA 2406 0601" or "FLO 2505 0114"
                        parts = str(bl_number).strip().split()
                        if len(parts) >= 2:
                            # Second part should be YYMM (4 digits)
                            potential_voy = ''.join(filter(str.isdigit, parts[1]))
                            if len(potential_voy) == 4:
                                logging.info(f"üìå Extracted VOY.NO. from BL number '{bl_number}': {potential_voy}")
                                return potential_voy

                    return 'Not Available'

                # Helper function to clean record
                def clean_record(record):
                    """Clean VOY.NO. field in a record"""
                    if isinstance(record, dict) and 'VOY.NO.' in record:
                        original_voy = record['VOY.NO.']
                        bl_number = record.get('BL number')
                        cleaned_voy = clean_voy_no(original_voy, bl_number)
                        if original_voy != cleaned_voy:
                            print(f"üßπ Cleaned VOY.NO.: '{original_voy}' ‚Üí '{cleaned_voy}' (BL: {bl_number})")
                        record['VOY.NO.'] = cleaned_voy
                    return record

                # Process each section
                for i, bl_section in enumerate(bl_sections, 1):
                    print(f"\n=== DEBUG: Processing BL Section {i}/{len(bl_sections)} ===")
                    print(f"Section length: {len(bl_section)} characters")
                    print(f"First 400 chars of section:\n{bl_section[:400]!r}")
                    print(f"analysis_type: {analysis_type}")
                    print(f"lang: {lang}")
                    if custom_prompt:
                        print(f"üß† custom_prompt (first 8 words): {' '.join(custom_prompt.split()[:8])}")

                    try:
                        result, tokens_used = analyze_text_with_prompt_with_gemini(
                            bl_section, analysis_type, lang, custom_prompt
                        )
                        # Add this logging
                        print(f"DEBUG ‚Üí LLM extracted for {filename}:")
                        if isinstance(result, list):
                            for item in result:
                                if isinstance(item, dict):
                                    bl_num = item.get('BL number', 'N/A')
                                    bl_qty = item.get('B/L quantity (MT)', 'N/A')
                                    split_qty = item.get('B/L split quantity (MT)', 'N/A')
                                    print(f"  BL: {bl_num} | B/L Qty: {bl_qty} | Split: {split_qty}")
                        total_tokens += tokens_used

                        print(f"DEBUG ‚Üí Tokens used: {tokens_used}")
                        print("DEBUG ‚Üí Raw result type:", type(result))
                        print("DEBUG ‚Üí Raw result value:", repr(result))

                        if isinstance(result, dict):
                            print("DEBUG ‚Üí Result keys:", list(result.keys()))

                    except Exception as e:
                        print(f"‚ö†Ô∏è ERROR while analyzing section {i}: {e}")
                        continue

                    # Skip if result is None or has error
                    if not result or (isinstance(result, dict) and result.get('error')):
                        print(f"‚ö†Ô∏è Skipping section {i} ‚Äî result is None or contains error.")
                        continue

                    print(f"‚úÖ Section {i} processed successfully.\n")

                    # Clean VOY.NO. in results
                    if isinstance(result, list):
                        result = [clean_record(item) for item in result]
                    elif isinstance(result, dict):
                        result = clean_record(result)

                    # Attach filename if requested
                    if keep_filename_param:
                        if isinstance(result, dict):
                            result['__filename__'] = filename
                        elif isinstance(result, list):
                            for item in result:
                                if isinstance(item, dict):
                                    item['__filename__'] = filename

                    # Flatten results into a single list
                    if isinstance(result, list):
                        all_results.extend(result)
                    else:
                        all_results.append(result)

                return all_results, total_tokens

            return llm_worker

        llm_worker = create_llm_worker(keep_filename)

        processed_files_llm = 0
        file_processing_logs = {}

        with ThreadPoolExecutor(max_workers=4) as executor:
            for idx, (result, tokens_used) in enumerate(executor.map(llm_worker, extracted)):
                processed_files_llm += 1
                filename = extracted[idx][0]
                file_path = extracted[idx][2]
                file_processing_logs[filename] = {
                    'llm_tokens': tokens_used,
                    'vlm_fields': [],
                    'vlm_tokens': 0,
                    'extracted_fields': 0,
                    'missing_fields': 0
                }
                log_entries.append(f"LLM Analysis: {filename}")
                log_entries.append(f"  - Tokens Used: {tokens_used}")
                log_entries.append(f"  - Analysis Result: {'Success' if result else 'Failed'}")
                update_progress(task_id, total_files + processed_files_llm, total_files * 2, f"Analyzing {filename}")
                app.logger.info(f"ARJ_in_llm_worker_2")

                if result:
                    if analysis_type == 'Cargo_BL':

                        def validate_bl_data(item):
                            """Validate B/L vs split quantity with detailed logging."""
                            bl_qty = item.get('B/L quantity (MT)')
                            split_qty = item.get('B/L split quantity (MT)')
                            bl_number = item.get('BL number', 'Unknown')

                            # Debug log
                            app.logger.info(f"üîç Validating BL {bl_number}: B/L qty='{bl_qty}', Split qty='{split_qty}'")

                            try:
                                # Clean and convert B/L quantity
                                if bl_qty not in [None, '', 'n/a', 'null']:
                                    bl_val = float(str(bl_qty).replace(',', '').strip())
                                else:
                                    bl_val = None

                                # Clean and convert split quantity
                                if split_qty not in [None, '', 'n/a', 'null']:
                                    split_val = float(str(split_qty).replace(',', '').strip())
                                else:
                                    split_val = None

                                app.logger.info(f"üîç Converted: bl_val={bl_val}, split_val={split_val}")

                                # Validation logic
                                if split_val is not None and bl_val is not None:
                                    # Check if split > B/L (invalid)
                                    if split_val > bl_val + 0.001:
                                        app.logger.warning(
                                            f"‚ùå INVALID: BL {bl_number} - Split ({split_val}) > B/L ({bl_val}). Clearing."
                                        )
                                        item['B/L split quantity (MT)'] = None

                                    # Check if split == B/L (suspicious - likely error)
                                    elif abs(split_val - bl_val) < 0.001:
                                        app.logger.warning(
                                            f"‚ö†Ô∏è EQUALS: BL {bl_number} - Split ({split_val}) == B/L ({bl_val}). Clearing."
                                        )
                                        item['B/L split quantity (MT)'] = None

                                    # Valid split quantity
                                    else:
                                        app.logger.info(f"‚úÖ Valid: BL {bl_number} - {split_val} <= {bl_val}")
                                        item['B/L split quantity (MT)'] = f"{split_val:,.3f}"

                                # Format quantities
                                if split_val is None or item.get('B/L split quantity (MT)') is None:
                                    item['B/L split quantity (MT)'] = None

                                if bl_val is not None:
                                    item['B/L quantity (MT)'] = f"{bl_val:,.3f}"
                                else:
                                    item['B/L quantity (MT)'] = None

                            except Exception as e:
                                app.logger.error(f"‚ùå Error validating BL {bl_number}: {e}")

                            return item

                        def fill_missing_fields(d):
                            app.logger.info(f"ARJ_in_llm_worker_fill_missing_fields")
                            nonlocal total_tokens
                            vlm_processed_fields = []
                            vlm_tokens_used = 0

                            # Log the BL number we're processing
                            bl_number = d.get('BL number', 'Unknown')
                            app.logger.info(f"üîç Processing BL: {bl_number}")

                            extracted_count = sum(1 for v in d.values() if not is_missing(v))
                            missing_count = sum(1 for v in d.values() if is_missing(v))
                            file_processing_logs[filename]['extracted_fields'] = extracted_count
                            file_processing_logs[filename]['missing_fields'] = missing_count

                            for key, value in d.items():
                                # Skip copy_type field
                                if key == 'copy_type':
                                    continue

                                # Log what we found for this field
                                if not is_missing(value):
                                    app.logger.info(f"  ‚úì {key}: '{value}' (already extracted)")
                                else:
                                    app.logger.info(f"  ‚úó {key}: Missing, will use VLM")

                                if is_missing(value):
                                    # Pass B/L quantity if we're filling split quantity
                                    bl_qty = d.get('B/L quantity (MT)') if key == 'B/L split quantity (MT)' else None

                                    # Log before VLM
                                    if bl_qty:
                                        app.logger.info(f"    ‚Üí VLM will validate against B/L qty: {bl_qty}")

                                    filled, vlm_tokens = gemini_vlm_field(
                                        file_path,
                                        key,
                                        lang,
                                        provider_override=vlm_provider,
                                        model_override=vlm_model,
                                        bl_quantity=bl_qty
                                    )

                                    # Log after VLM
                                    app.logger.info(f"    ‚Üí VLM returned: '{filled}'")

                                    d[key] = filled
                                    total_tokens += vlm_tokens
                                    vlm_tokens_used += vlm_tokens
                                    vlm_processed_fields.append(f"{key}={filled}")
                                    logging.info(
                                        f"VLM Ë£úÊïë: Ê™îÊ°à={os.path.basename(file_path)}, Ê¨Ñ‰Ωç={key}, Ë£úÊïëÂÄº={filled}")

                            file_processing_logs[filename]['vlm_fields'] = vlm_processed_fields
                            file_processing_logs[filename]['vlm_tokens'] = vlm_tokens_used
                            return d

                        def fill_missing_charterers_by_cargo(data_list):
                            """Fill missing charterers based on same Cargo # having same charterer"""

                            # Group by Cargo #
                            cargo_groups = {}
                            for item in data_list:
                                cargo_num = item.get('Cargo #')
                                if cargo_num:
                                    if cargo_num not in cargo_groups:
                                        cargo_groups[cargo_num] = []
                                    cargo_groups[cargo_num].append(item)

                            # For each cargo group, find the charterer and apply to all
                            for cargo_num, items in cargo_groups.items():
                                # Find first valid charterer in this group
                                valid_charterer = None
                                for item in items:
                                    charterer = item.get('Charterer')
                                    if charterer and charterer not in ['Not Available', 'null', None, '']:
                                        valid_charterer = charterer
                                        break

                                # Apply to all items in this cargo group
                                if valid_charterer:
                                    for item in items:
                                        current_charterer = item.get('Charterer')
                                        if not current_charterer or current_charterer == 'Not Available':
                                            item['Charterer'] = valid_charterer
                                            app.logger.info(
                                                f"üîÑ Filled Charterer for Cargo #{cargo_num}, BL {item.get('BL number')}: '{valid_charterer}'")

                            return data_list

                    if isinstance(result, dict) and 'data' in result:
                        cargo_data = result.get('data', [])
                        if isinstance(cargo_data, list):
                            for d in cargo_data:
                                fill_missing_fields(d)
                                validate_bl_data(d)
                                all_data.append(d)
                    elif isinstance(result, list):
                        for d in result:
                            fill_missing_fields(d)
                            validate_bl_data(d)
                            all_data.append(d)
                    elif isinstance(result, dict):
                        fill_missing_fields(result)
                        validate_bl_data(result)
                        all_data.append(result)
                    else:
                        if isinstance(result, dict):
                            all_data.append(result)
                        elif isinstance(result, list):
                            all_data.extend(result)

                    total_tokens += tokens_used

                    # After processing all files, fill missing charterers by cargo group
                    all_data = fill_missing_charterers_by_cargo(all_data)

        # Postprocessing
        import re
        import pprint
        def clean_and_convert_qty(value):
            """
            Robustly cleans quantity-related strings by removing common units,
            handles both European (5.349.878) and US (5,349.878) formats.
            """
            if value is None or value in ['', 'null', 'n/a']:
                return None

            # Convert to string
            text = str(value)

            # Remove common units
            text = text.replace('MTS', '').replace('MT', '').replace('tons', '').replace('Âô∏', '').strip()

            # Count dots and commas to determine format
            dot_count = text.count('.')
            comma_count = text.count(',')

            # European format with multiple dots (5.349.878)
            if dot_count > 1:
                # Remove thousand separators, keep last dot as decimal
                parts = text.split('.')
                if len(parts) > 1:
                    text = ''.join(parts[:-1]) + '.' + parts[-1]
            # US format with commas (5,349.878 or 5,349,878)
            elif comma_count >= 1:
                text = text.replace(',', '')

            # Extract the number
            match = re.search(r'([\d\.]+)', text)
            if match:
                try:
                    return float(match.group(1))
                except (ValueError, TypeError):
                    return None
            return None

        if all_data and analysis_type == 'Cargo_BL':
            log_entries = []  # Ensure log_entries is initialized

            # --- SINGLE, UNIFIED PROCESSING LOOP ---
            # All data cleaning and validation happens here, once per item.
            for item in all_data:
                # 1. Apply OCR fixes to all string values first
                for key, value in item.items():
                    if isinstance(value, str):
                        item[key] = fix_ocr_errors(value)

                # 2. Clean and convert B/L and Split quantities
                bl_val = clean_and_convert_qty(item.get('B/L quantity (MT)'))
                split_val = clean_and_convert_qty(item.get('B/L split quantity (MT)'))

                # 3. Apply validation logic for Split Quantity
                if split_val is not None and bl_val is not None:
                    # Use a small tolerance for float comparison
                    if split_val > bl_val + 0.001:
                        # ‚ùå Split > B/L ‚Üí Invalid, set to None
                        item['B/L split quantity (MT)'] = None
                        log_entries.append(
                            f"SID [WARN] Split qty > B/L qty for BL={item.get('BL number', 'N/A')} | "
                            f"Split={split_val} > B/L={bl_val} ‚Üí cleared."
                        )
                    elif abs(split_val - bl_val) < 0.001:
                        # ‚ö†Ô∏è Split == B/L ‚Üí Suspicious (likely no actual split), set to None
                        item['B/L split quantity (MT)'] = None
                        log_entries.append(
                            f"SID [WARN] Split qty equals B/L qty for BL={item.get('BL number', 'N/A')} | "
                            f"Split={split_val} == B/L={bl_val} ‚Üí cleared (no actual split)."
                        )
                    else:
                        # ‚úÖ Valid split: split < B/L
                        item['B/L split quantity (MT)'] = f"{split_val:,.3f}"
                        log_entries.append(
                            f"SID [INFO] Valid split qty for BL={item.get('BL number', 'N/A')} | "
                            f"Split={split_val} < B/L={bl_val} ‚Üí kept."
                        )
                elif split_val is not None:
                    # Split exists but no B/L quantity to compare against (unusual)
                    item['B/L split quantity (MT)'] = f"{split_val:,.3f}"
                else:
                    # No split quantity found
                    item['B/L split quantity (MT)'] = None

                # 4. Format B/L quantity with commas
                if bl_val is not None:
                    item['B/L quantity (MT)'] = f"{bl_val:,.3f}"
                else:
                    item['B/L quantity (MT)'] = None

                # 4. Format the final B/L Quantity with commas
                if bl_val is not None:
                    item['B/L quantity (MT)'] = f"{bl_val:,.3f}"  # Add comma here
                else:
                    item['B/L quantity (MT)'] = None

                # 5. Validate OBL release date format
                obl_date = item.get('OBL release date')
                if not (isinstance(obl_date, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', obl_date)):
                    item['OBL release date'] = None

            # --- FINAL DEBUG LOGGING (runs ONCE after all processing is complete) ---
            #app.logger.info(f"Post-processing completed for task {task_id}: {len(all_data)} items processed.")
            pp = pprint.PrettyPrinter(depth=4)
#            for i, final_item in enumerate(all_data, 1):
 #               app.logger.info(f"Final Row {i}: {pp.pformat(final_item)}")
            app.logger.info("=== DEBUG: all_data BEFORE cargo_bl_postprocess ===")
            app.logger.info("=== DEBUG: all_data BEFORE cargo_bl_postprocess ===")
            if all_data:
                df = pd.DataFrame(all_data)
                app.logger.info("\n" + tabulate(df, headers='keys', tablefmt='psql', showindex=True))
            else:
                app.logger.info("‚ö†Ô∏è all_data is empty before cargo_bl_postprocess.")

            # ALWAYS call cargo_bl_postprocess (moved outside the if/else)
            from analyzers.services import cargo_bl_postprocess
            all_data = cargo_bl_postprocess(all_data, keep_filename)

            # Debug: Check after calling cargo_bl_postprocess
            has_filename_after_postprocess = any('File name' in item for item in all_data if isinstance(item, dict))
            if has_filename_after_postprocess:
                sample_item = next((item for item in all_data if isinstance(item, dict) and 'File name' in item), None)
                if sample_item:
                    app.logger.info(f"Sample item with 'File name': {sample_item.get('File name')}")
            # Debug: Check after calling cargo_bl_postprocess
            has_filename_after_postprocess = any('File name' in item for item in all_data if isinstance(item, dict))
            #app.logger.info(f"After cargo_bl_postprocess: has 'File name' = {has_filename_after_postprocess}")
            if has_filename_after_postprocess:
                sample_item = next((item for item in all_data if isinstance(item, dict) and 'File name' in item), None)
                if sample_item:
                    app.logger.info(f"Sample item with 'File name': {sample_item.get('File name')}")
        # Áç≤ÂèñÂàóÂêçÔºàcargo_bl_postprocess Â∑≤Á∂ìËôïÁêÜ‰∫Ü __filename__ Â≠óÊÆµÔºâ
            headers = list(all_data[0].keys()) if all_data else []

            # Ë™øË©¶Êó•Ë™å
            app.logger.info(f"keep_filename: {keep_filename}")
            app.logger.info(f"Final headers: {headers}")
            app.logger.info(f"File name in headers: {'File name' in headers}")

            rows = []
            for i, item in enumerate(all_data):
                row = []
                for header in headers:
                    value = item.get(header, '')
                    row.append(str(value) if value is not None else '')
                is_total_row = i == len(all_data) - 2
                is_cgos_row = i == len(all_data) - 1
                rows.append({
                    'data': row,
                    'is_total': is_total_row,
                    'is_cgos': is_cgos_row
                })
            excel_filename = f"{analysis_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Áç≤ÂèñÁï∂ÂâçÁî®Êà∂‰ø°ÊÅØ
            current_user = get_user_by_username(username)
            user_info = {
                'display_name': current_user.get('display_name') if current_user else None,
                'logo_file': current_user.get('logo_file') if current_user else None
            }

            # Ê™¢Êü•ÊòØÂê¶Êúâ VESSEL Âíå VOY.NO. ‰ø°ÊÅØÔºàÈáùÂ∞ç Cargo_BLÔºâ
            vessel_voy_info = globals().get('vessel_voy_data', None)
            # === DEBUG: Log all_data before Excel export ===
            pp = pprint.PrettyPrinter(depth=4)
            app.logger.info("=== DEBUG: all_data BEFORE export_to_excel ===")
            if all_data:
                # Remove copy_type column if it still exists
                for item in all_data:
                    if 'copy_type' in item:
                        del item['copy_type']

                df = pd.DataFrame(all_data)
                # Limit columns if there are too many
                app.logger.info("\n" + tabulate(df, headers='keys', tablefmt='psql', showindex=True))
            else:
                app.logger.info("‚ö†Ô∏è all_data is empty.")
            excel_path = export_to_excel(all_data, excel_filename, 'Hansa Tankers', user_info=user_info,
                                         include_logo=True, vessel_voy_info=vessel_voy_info,
                                         keep_filename=keep_filename)
            # ÁîüÊàêÂ∞çÊáâÁöÑ log Ê™îÊ°àÔºå‰øùÂ≠òÂú® logs Ë≥áÊñôÂ§æ‰∏≠
            log_filename = excel_filename.replace('.xlsx', '_log.txt')

            # Á¢∫‰øù logs Ë≥áÊñôÂ§æÂ≠òÂú®
            logs_dir = 'logs'
            if not os.path.exists(logs_dir):
                os.makedirs(logs_dir)

            # Define log path and time cost
            log_path = os.path.join(logs_dir, log_filename)
            time_cost = time.time() - start_time

            # === ADD MORE DETAILED SUMMARY BEFORE END TIME ===
            log_entries.append("")
            log_entries.append("=== Processing Summary ===")
            log_entries.append(f"Total Processing Time: {time_cost:.2f} seconds")
            log_entries.append(f"Total Tokens Used: {total_tokens}")
            log_entries.append(f"Total Pages: {total_pages}")
            log_entries.append(f"Successfully Processed Files: {len(extracted)}")
            log_entries.append("")

            # === ADD DETAILED FILE PROCESSING STATISTICS ===
            log_entries.append("=== Detailed File Processing Statistics ===")
            for fname, flog in file_processing_logs.items():
                log_entries.append(f"File: {fname}")
                log_entries.append(f"  - LLM Tokens: {flog['llm_tokens']}")
                log_entries.append(f"  - VLM Tokens: {flog['vlm_tokens']}")
                log_entries.append(f"  - Total Tokens: {flog['llm_tokens'] + flog['vlm_tokens']}")
                log_entries.append(f"  - Successfully Extracted Fields: {flog['extracted_fields']}")
                log_entries.append(f"  - Missing Fields: {flog['missing_fields']}")
                if flog['vlm_fields']:
                    log_entries.append(f"  - VLM Remediated Fields: {', '.join(flog['vlm_fields'])}")
                log_entries.append("")

            # === FINAL SUMMARY ===
            log_entries.append(f"End Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log_entries.append(f"Output Excel: {excel_filename}")
            log_entries.append(f"Output Log: {log_filename}")

            # === WRITE LOG FILE ===
            with open(log_path, 'w', encoding='utf-8') as log_file:
                log_file.write('\n'.join(log_entries))

            logging.info(f"Detailed log file generated: {log_filename}")

            # Optional: rounded total time for final display or print
            time_cost = round(time.time() - start_time, 2)

            # Áõ¥Êé•ÁîüÊàêËã±ÊñáÊó•Ë™å
            with open(log_path, 'w', encoding='utf-8') as log_file:
                log_file.write('\n'.join(log_entries))
            logging.info(f"Log file generated: {log_filename}")

            time_cost = round(time.time() - start_time, 2)

            # Ê†πÊìöË®≠ÁΩÆÊ±∫ÂÆöÊòØÂê¶‰øùÂ≠òÊñá‰ª∂
            saved_files = []
            if save_files:
                # ‰øùÂ≠òÊñá‰ª∂ÔºåË®òÈåÑÊñá‰ª∂Ë∑ØÂæëÂà∞Ê≠∑Âè≤Ë®òÈåÑ‰∏≠
                for filename, _, file_path in extracted:
                    if os.path.exists(file_path):
                        # ÈáçÂëΩÂêçÊñá‰ª∂ÁÇ∫Êõ¥ÂèãÂ•ΩÁöÑÂêçÁ®±
                        file_ext = os.path.splitext(filename)[1]
                        new_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(filename)}"
                        new_file_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
                        try:
                            os.rename(file_path, new_file_path)
                            saved_files.append((filename, new_filename))
                            log_entries.append(f"File Saved: {filename} -> {new_filename}")
                        except Exception as e:
                            app.logger.warning(f"Failed to rename saved file: {file_path} {e}")
                            # Â¶ÇÊûúÈáçÂëΩÂêçÂ§±ÊïóÔºå‰øùÊåÅÂéüÊñá‰ª∂Âêç
                            saved_files.append((filename, os.path.basename(file_path)))
            else:
                # Áµ±‰∏ÄÂà™Èô§ÊâÄÊúâËá®ÊôÇÊ™îÊ°à
                for _, _, file_path in extracted:
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    except Exception as e:
                        app.logger.warning(f"Failed to delete temporary file: {file_path} {e}")

            history_entry = {
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'files': [f['filename'] for f in file_data_list],
                'analysis_type': analysis_type,
                'lang': lang,
                'rows': len(all_data),
                'cols': len(headers),
                'headers': headers,
                'tokens': total_tokens,
                'excel': excel_filename,
                'log_file': log_filename,  # Êñ∞Â¢û log Ê™îÊ°àË®òÈåÑ
                'time_cost': time_cost,
                'duration_str': f"{time_cost:.2f} sec.",
                'username': username,
                'total_pages': total_pages,  # Êñ∞Â¢ûÈ†ÅÊï∏Ë®òÈåÑ
                'saved_files': saved_files  # Ë®≠ÁΩÆ‰øùÂ≠òÁöÑÊñá‰ª∂ÂàóË°®
            }
            save_to_history(history_entry)
            app.logger.info(f"About to set final progress for task {task_id}")
            update_progress(task_id, total_files * 2, total_files * 2, 'Analysis completed!')
            app.logger.info(f"Updated final progress for task {task_id}")

            # ÊèêÂèñÁ¥îÊï∏ÊìöÁî®ÊñºÂâçÁ´ØÈ°ØÁ§∫
            simple_rows = [row['data'] if isinstance(row, dict) and 'data' in row else row for row in rows]
            app.logger.info(f"About to set final_result for task {task_id}: prepared {len(simple_rows)} rows")

            with progress_lock:
                progress_store[task_id]['final_result'] = {
                    'success': True,
                    'headers': headers,
                    'rows': simple_rows,
                    'excel': excel_filename,
                    'log_file': log_filename,  # Êñ∞Â¢û log Ê™îÊ°àË®òÈåÑ
                    'lang_code': lang,
                    'total_pages': total_pages
                }
                app.logger.info(f"Successfully set final_result for task {task_id}: success=True, rows={len(simple_rows)}, excel={excel_filename}")
        else:
            app.logger.warning(f"No data extracted for task {task_id}, setting failure result")
            with progress_lock:
                progress_store[task_id]['final_result'] = {
                    'success': False,
                    'error': 'Unable to extract valid data from files'
                }
                app.logger.info(f"Set final_result for task {task_id}: success=False (no data)")
        # Â∞á‰øùÂ≠òÁöÑÊñá‰ª∂‰ø°ÊÅØÊ∑ªÂä†Âà∞ÈÄ≤Â∫¶Â≠òÂÑ≤‰∏≠
        if saved_files:
            with progress_lock:
                progress_store[task_id]['saved_files'] = saved_files
    except Exception as e:
        app.logger.error(f"Error occurred during processing: {str(e)}", exc_info=True)
        with progress_lock:
            progress_store[task_id]['final_result'] = {
                'success': False,
                'error': str(e)
            }
            app.logger.info(f"Set final_result for task {task_id}: success=False due to exception: {str(e)}")

@app.route('/progress/<task_id>')
def check_progress(task_id):
    """Check task progress"""
    return get_progress(task_id)

@app.route('/export/<filename>')
def export(filename):
    # First check uploads directory
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(upload_path):
        return send_file(upload_path, as_attachment=True, download_name=filename)
    
    # Then check temp directory
    temp_path = os.path.join(tempfile.gettempdir(), filename)
    if os.path.exists(temp_path):
        return send_file(temp_path, as_attachment=True, download_name=filename)
    
    return "File not found", 404

@app.route('/logs/<filename>')
def download_log(filename):
    """Download log file"""
    # Check logs directory
    logs_path = os.path.join('logs', filename)
    if os.path.exists(logs_path):
        return send_file(logs_path, as_attachment=True, download_name=filename)
    
    return "Log file not found", 404

@app.route('/saved_files/<filename>')
def download_saved_file(filename):
    """Download saved analysis file"""
    # Check uploads directory for saved files
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    
    return "File not found", 404




@app.route('/download_zip/<int:index>')
def download_zip_files(index):
    """Download all saved files as ZIP for a specific analysis record"""
    try:
        # Align index with admin dashboard ordering (time desc)
        def _get_history_time(x):
            return x.get('time') or x.get('timestamp') or ''
        sorted_history = sorted(history, key=_get_history_time, reverse=True)

        # Get the history record by index (from sorted list)
        if index < 0 or index >= len(sorted_history):
            app.logger.error(f"History index {index} not found, total history: {len(sorted_history)} (sorted)")
            return "Analysis record not found", 404

        record = sorted_history[index]
        saved_files = record.get('saved_files', [])
        
        app.logger.info(f"ZIP download request for index {index}, saved_files: {saved_files}")
        
        if not saved_files:
            app.logger.warning(f"No saved files for index {index}")
            return "No saved files found for this analysis record", 404
        
        # Create ZIP file in uploads directory (same as other files)
        import zipfile
        from datetime import datetime
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f"analysis_files_{index}_{timestamp}.zip"
        zip_path = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)
        
        app.logger.info(f"Creating ZIP file at: {zip_path}")
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            files_added = 0
            for original_name, saved_name in saved_files:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_name)
                app.logger.info(f"Checking file: {file_path}")
                
                if os.path.exists(file_path):
                    # Add file to ZIP with original name
                    zipf.write(file_path, original_name)
                    files_added += 1
                    app.logger.info(f"Added {original_name} to ZIP from {file_path}")
                else:
                    app.logger.warning(f"File not found: {file_path}")
        
        if files_added == 0:
            if os.path.exists(zip_path):
                os.remove(zip_path)
            app.logger.error("No files could be added to ZIP")
            return "No files could be added to ZIP", 404
        
        app.logger.info(f"ZIP file created successfully with {files_added} files")
        
        # Use the same route as other downloads
        return send_file(
            zip_path,
            as_attachment=True,
            download_name=f"analysis_files_{record.get('time', 'unknown').replace(':', '-').replace(' ', '_')}.zip"
        )
        
    except Exception as e:
        app.logger.error(f"Error creating ZIP file: {e}", exc_info=True)
        return f"Error creating ZIP file: {str(e)}", 500

@app.route('/history')
@login_required
def get_history():
    lang = 'en'  # Default to English
    # Only show own history: will take effect after username is added to history entry; allow all for now, add username when saving later
    def get_history_time(x):
        return x.get('time') or x.get('timestamp') or ''
    # Safe filtering: All users (including administrators) can only see their own records on /history page
    # Administrators should go to /admin page to see all records
    username = session.get('username')
    
    # All users can only see their own records
    filtered = [h for h in history if h.get('username') == username]
    
    sorted_history = sorted(filtered, key=get_history_time, reverse=True)
    return render_template('history.html', history=sorted_history, config=config, lang=lang)

@app.route('/clear_history', methods=['POST'])
def clear_history():
    # ÂÖàÂà™Èô§ÊâÄÊúâÂ∑≤Â≠òÁöÑ Excel Ê™îÊ°à
    for h in history:
        excel_name = h.get('excel')
        if excel_name:
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_name)
            try:
                if os.path.exists(excel_path):
                    os.remove(excel_path)
            except Exception:
                pass
    history.clear()
    save_history()
    return jsonify({'ok': True})

def fix_ocr_errors(text):
    """Fix common OCR errors"""
    # Fix Row -> Stow
    text = re.sub(r'\bRow\b', 'Stow', text, flags=re.IGNORECASE)
    # Can add more OCR error correction rules
    return text


def enforce_column_order_list(data_list):
    """Enforce column order for list data"""
    if not data_list:
        return data_list
    
    desired_order = [
        'Cargo #', 'BL number', 'B/L quantity (MT)', 'B/L split quantity (MT)',
        'Cargo name', 'Charterer', 'Consignee, order to', 'Notify',
        'Stow', 'LoadPort', 'Disch. Port', 'OBL release date', 'Release cargo against'
    ]
    
    # Reorder each item's columns
    ordered_data = []
    for item in data_list:
        ordered_item = {}
        
        # First add columns in desired_order
        for col in desired_order:
            if col in item:
                ordered_item[col] = item[col]
        
        # Then add other columns
        for key, value in item.items():
            if key not in ordered_item:
                ordered_item[key] = value
        
        ordered_data.append(ordered_item)
    
    return ordered_data

def enforce_column_order(df, desired_order):
    """Reorder DataFrame columns according to specified order"""
    # Get existing columns
    existing_cols = df.columns.tolist()
    
    # Create new column order
    ordered_cols = []
    
    # First add columns that exist in desired_order
    for col in desired_order:
        if col in existing_cols:
            ordered_cols.append(col)
    
    # Then add other columns not in desired_order
    for col in existing_cols:
        if col not in ordered_cols:
            ordered_cols.append(col)
    
    return df[ordered_cols]

def sort_by_bl_number(df):
    """Sort by BL number"""
    if 'BL number' in df.columns:
        # Remove empty values and sort
        df_sorted = df.dropna(subset=['BL number']).sort_values('BL number')
        df_empty = df[df['BL number'].isna()]
        return pd.concat([df_sorted, df_empty], ignore_index=True)
    return df

def post_process_cargo_bl_data(df):
    """Post-process Cargo_BL data, handle special rules"""
    if 'Release cargo against' in df.columns:
        # Replace N/A with LOI
        df['Release cargo against'] = df['Release cargo against'].replace('N/A', 'LOI')
        df['Release cargo against'] = df['Release cargo against'].fillna('LOI')

    # Add total row
    if len(df) > 0:
        # 1. Total row: Calculate sum of B/L quantity and B/L split quantity
        total_row = {}
        for col in df.columns:
            if col == 'Cargo #':
                total_row[col] = 'Total'
            elif col == 'B/L quantity (MT)':
                # Calculate sum, handle possible string format (e.g., "1,007.659")
                total_sum = 0
                for val in df[col].dropna():
                    try:
                        # Remove commas and convert to float
                        clean_val = str(val).replace(',', '').strip()
                        if clean_val and clean_val != 'nan':
                            total_sum += float(clean_val)
                    except (ValueError, TypeError):
                        continue
                total_row[col] = f"{total_sum:,.3f}"
            elif col == 'B/L split quantity (MT)':
                # Calculate sum
                total_sum = 0
                for val in df[col].dropna():
                    try:
                        clean_val = str(val).replace(',', '').strip()
                        if clean_val and clean_val != 'nan':
                            total_sum += float(clean_val)
                    except (ValueError, TypeError):
                        continue
                total_row[col] = f"{total_sum:,.3f}"
            else:
                total_row[col] = ''

        # 2. # Cgos Ë°åÔºöË®àÁÆó B/L number ÁöÑÊï∏Èáè
        cgos_row = {}
        bl_count = df['BL number'].dropna().nunique() if 'BL number' in df.columns else 0

        for col in df.columns:
            if col == 'Cargo #':
                cgos_row[col] = f'# Cgos: {bl_count}'
            else:
                cgos_row[col] = ''

        # Â∞áÁ∏ΩË®àË°åÊ∑ªÂä†Âà∞ DataFrame
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        df = pd.concat([df, pd.DataFrame([cgos_row])], ignore_index=True)

    return df

def safe_float(val):
    try:
        if isinstance(val, str):
            val = val.replace(',', '')
        return float(val)
    except Exception:
        return 0.0

def process_same_category_bl(data_list):
    """
    ËôïÁêÜÂêåÈ°û B/L ÁöÑ B/L quantity Âíå Cargo# ÈáçË§áÂïèÈ°å
    Ê†πÊìö BL number ÁöÑÁ¨¨10-11‰ΩçÁõ∏Âêå‰ΩÜÁ¨¨12-13‰Ωç‰∏çÂêå‰æÜÂà§Êñ∑ÂêåÈ°û
    
    ‰æãÂ¶ÇÔºö
    - "EMA 2406 0801" Âíå "EMA 2406 0802" 
    - Á¨¨10-11‰ΩçÈÉΩÊòØ "08"ÔºåÁ¨¨12-13‰ΩçÂàÜÂà•ÊòØ "01" Âíå "02"
    - ÂÆÉÂÄëÂ±¨ÊñºÂêå‰∏ÄÈ°ûÔºåÂè™ÊúâÁ¨¨‰∏ÄÂÄã‰øùÁïô B/L quantity Âíå Cargo#
    """
    if not data_list:
        return data_list
    
    # Êåâ BL number ÂàÜÁµÑ
    bl_groups = {}
    
    for i, item in enumerate(data_list):
        bl_number = item.get('BL number', '').strip()
        if not bl_number:
            continue
            
        # ÊèêÂèñ BL number ÁöÑÁ¨¨10-11‰ΩçÁî®ÊñºÂàÜÁµÑ
        # Ê†ºÂºèÔºöEMA 2406 0804
        #       123456789012345
        # Á¨¨10-11‰ΩçÊòØÁ¨¨‰∏âÈÉ®ÂàÜÁöÑÂâçÂÖ©‰Ωç
        if len(bl_number) >= 13:
            try:
                # Áõ¥Êé•ÂèñÁ¨¨10-11‰ΩçÂ≠óÁ¨¶
                group_key = bl_number[9:11]  # Á¨¨10-11‰Ωç (0-based index)
                
                if group_key not in bl_groups:
                    bl_groups[group_key] = []
                bl_groups[group_key].append((i, item))
            except (IndexError, ValueError):
                # Â¶ÇÊûú BL number Ê†ºÂºè‰∏çÁ¨¶ÂêàÈ†êÊúüÔºåÂñÆÁç®ËôïÁêÜ
                continue

    # ËôïÁêÜÊØèÂÄãÂàÜÁµÑ
    for group_key, group_items in bl_groups.items():
        print(f"\nüîπ Processing group: {group_key} (items: {len(group_items)})")

        if len(group_items) > 1:  # Âè™ËôïÁêÜÊúâÂ§öÂÄãÈ†ÖÁõÆÁöÑÂàÜÁµÑ
            # Êåâ BL number ÊéíÂ∫èÔºåÁ¢∫‰øùËôïÁêÜÈ†ÜÂ∫è‰∏ÄËá¥
            group_items.sort(key=lambda x: x[1].get('BL number', ''))
            print("  ‚Ü≥ Sorted items by BL number.")

            prev_bl_number = None  # Áî®ÊñºË®òÈåÑÂâç‰∏ÄÂÄã BL number

            # ÈÄê‰∏ÄËôïÁêÜÈ†ÖÁõÆ
            for idx, (original_index, item) in enumerate(group_items):
                current_bl = item.get('BL number', '')
                print(f"    ‚Ä¢ Item {idx + 1} (original index {original_index}): BL={current_bl}")

                if idx == 0:
                    # Á¨¨‰∏ÄÂÄãÈ†ÖÁõÆ‰øùÊåÅ‰∏çËÆä
                    print("      ‚Üí First item in group: keeping original values.")
                    prev_bl_number = current_bl
                    continue

                # Ê∏ÖÁ©∫ Cargo #
                if 'Cargo #' in item:
                    item['Cargo #'] = None
                    print("      ‚Üí Cleared Cargo #")

                # Ê†πÊìö BL number Ê±∫ÂÆöÊòØÂê¶Ê∏ÖÁ©∫ B/L quantity
                if current_bl == prev_bl_number:
                    if 'B/L quantity (MT)' in item:
                        item['B/L quantity (MT)'] = None
                    print("      ‚Üí Same BL as previous, cleared B/L quantity (MT).")
                else:
                    # Â¶ÇÊûú‰∏çÂêåÔºå‰øùÁïôÂéüÂßãÊï∏ÂÄº
                    print("      ‚Üí New BL number detected, kept B/L quantity (MT).")
                    prev_bl_number = current_bl

        else:
            print(f"  ‚ö™ Group has only one item ‚Äî skipped.")
    return data_list

def process_single_file(file_data, analysis_type, prompt_template, keep_filename):
    """ËôïÁêÜÂñÆ‰∏ÄÊ™îÊ°àÁöÑÂáΩÊï∏ÔºåÁî®Êñº‰∏¶Ë°åËôïÁêÜ"""
    file, filename = file_data
    try:
        suffix = filename.split('.')[-1].lower()
        
        # Ê™¢Êü•Ê™îÊ°àÂ§ßÂ∞è
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        
        if file_size == 0:
            print(f"ÈåØË™§ÔºöÊ™îÊ°àÂ§ßÂ∞èÁÇ∫ 0 - {filename}")
            return None, 0
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.'+suffix) as tmp:
            file.seek(0)
            file.save(tmp.name)
            
            saved_size = os.path.getsize(tmp.name)
            if saved_size == 0:
                file.seek(0)
                content = file.read()
                if len(content) > 0:
                    with open(tmp.name, 'wb') as f:
                        f.write(content)
                    saved_size = os.path.getsize(tmp.name)
                else:
                    return None, 0
            
            # ÊñáÂ≠óÊì∑Âèñ
            text = extract_text(tmp.name, suffix)
            # OCR ÈåØË™§‰øÆÊ≠£
            text = fix_ocr_errors(text)
            
            # Ê∏ÖÁêÜËá®ÊôÇÊ™îÊ°à
            try:
                os.unlink(tmp.name)
            except:
                pass
            
            if len(text.strip()) < 10:  # ÈÅéÊøæÂ§™Áü≠ÁöÑÊñáÂ≠ó
                print(f"Ë≠¶ÂëäÔºöÊ™îÊ°à {filename} Ëß£ÊûêÂá∫ÁöÑÊñáÂ≠óÂ§™Áü≠ÔºåË∑≥ÈÅé")
                return None, 0
            
            # Ê∫ñÂÇô prompt
            if '{text}' not in prompt_template:
                prompt_str = f"{prompt_template}\n\nÊ™îÊ°àÂÖßÂÆπÔºö\n{text}"
            else:
                prompt_str = prompt_template.replace('{text}', text)
            
            # ÂëºÂè´ API
            result_json, tokens_used = call_gemini_api(prompt_str)
            data = extract_json_from_response(result_json)
            
            if keep_filename:
                if isinstance(data, dict):
                    data['__filename__'] = filename
                elif isinstance(data, list):
                    for d in data:
                        if isinstance(d, dict):
                            d['__filename__'] = filename
            
            print(f"‚úì ÂÆåÊàêËôïÁêÜÔºö{filename} (Token: {tokens_used})")
            return data, tokens_used
            
    except Exception as e:
        print(f"Ê™îÊ°àËôïÁêÜÈåØË™§ {filename}Ôºö{e}")
        return None, 0

@app.route('/export_custom', methods=['POST'])
def export_custom():
    try:
        data = request.get_json()
        headers = data.get('headers')
        rows = data.get('rows')
        manual_data = data.get('manual_data', {'headers': [], 'rows': []})
        include_logo = data.get('include_logo', False)
        preview_title = data.get('preview_title', 'Hansa Tankers')
        
        if not headers or not rows:
            return jsonify({'success': False, 'error': 'Áº∫Â∞ëË≥áÊñô'})
        
        import pandas as pd
        # Áî¢ÁîüÂîØ‰∏ÄÊ™îÂêç
        filename = f"CustomExport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Áç≤ÂèñÁï∂ÂâçÁî®Êà∂‰ø°ÊÅØ  
        current_user = get_user_by_username(session.get('username'))
        user_info = {
            'display_name': current_user.get('display_name') if current_user else None,
            'logo_file': current_user.get('logo_file') if current_user else None
        }
        
        # Áõ¥Êé•ÂëºÂè´ export_to_excelÔºåÂ∏∂ÂÖ•ÊâÄÊúâÊñ∞ÂèÉÊï∏
        excel_path = export_to_excel(
            [dict(zip(headers, row)) for row in rows], 
            filename, 
            preview_title,
            manual_data=manual_data,
            include_logo=include_logo,
            user_info=user_info
        )
        
        if excel_path is None:
            return jsonify({'success': False, 'error': 'Failed to create Excel file'})
        
        # Ê™¢Êü•Ê™îÊ°àÊòØÂê¶ÁúüÁöÑÂ≠òÂú®
        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': f'Excel file not found: {excel_path}'})
        
        # Ê≥®ÊÑèÔºöÈÄôË£°‰∏çÂÜçÂÑ≤Â≠òÊ≠∑Âè≤Ë®òÈåÑÔºåÂõ†ÁÇ∫ÈÄôÂè™ÊòØÈáçÊñ∞ÂåØÂá∫Â∑≤ÂàÜÊûêÁöÑË≥áÊñô
        return jsonify({'success': True, 'download_url': url_for('download_custom_excel', filename=filename)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download_custom_excel/<filename>')
def download_custom_excel(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

def is_missing(value):
    return value is None or str(value).strip() == '' or str(value).lower() in ['null', 'n/a']


def gemini_vlm_field(file_path, field_name, lang='zh', provider_override: Optional[str] = None,
                     model_override: Optional[str] = None, bl_quantity: Optional[str] = None):
    # Ê†πÊìöÂëºÂè´ÊñπÊåáÂÆöÊàñ‰ΩøÁî®ËÄÖË®≠ÂÆöÊ±∫ÂÆö‰ΩøÁî® cloud Êàñ local(VLM: ollama)
    provider = provider_override
    ollama_model = model_override
    if provider is None:
        user = get_current_user()
        if user:
            setting = get_user_analyzer_setting(user['id'], 'Cargo_BL')
            provider = setting.get('vlm_provider', 'cloud')
            ollama_model = ollama_model or setting.get('ollama_model')
        else:
            provider = 'cloud'

    # Helper function to clean VOY.NO.
    def clean_voy_no(voy_value):
        """Extract only 4-digit voyage numbers, return 'Not Available' if invalid"""
        if not voy_value:
            return 'Not Available'

        # Check if value already indicates "not found"
        if any(phrase in str(voy_value).lower() for phrase in ['not found', 'not available', 'n/a', 'null', 'Êâæ‰∏çÂà∞']):
            return 'Not Available'

        # Extract digits
        cleaned = ''.join(filter(str.isdigit, str(voy_value)))

        # VOY.NO. must be exactly 4 digits (YYMM format like 2406, 2505)
        if len(cleaned) == 4:
            return cleaned
        else:
            # Invalid format (too short, too long, or empty)
            logging.info(
                f"‚ö†Ô∏è Invalid VOY.NO. format: '{voy_value}' (cleaned: '{cleaned}', length: {len(cleaned)}) ‚Üí 'Not Available'")
            return 'Not Available'

    # Helper function to validate split quantity
    def validate_split_qty(split_value, bl_value):
        """Validate split quantity against B/L quantity"""
        if not split_value or not bl_value:
            return split_value

        try:
            # Clean the values
            split_str = str(split_value).replace(',', '').replace('MT', '').replace('MTS', '').strip()
            bl_str = str(bl_value).replace(',', '').replace('MT', '').replace('MTS', '').strip()

            split_float = float(split_str)
            bl_float = float(bl_str)

            # If split > BL, invalid
            if split_float > bl_float + 0.001:
                logging.warning(f"‚ö†Ô∏è VLM extracted invalid split qty: {split_value} > B/L: {bl_value} ‚Üí returning None")
                return None

            # If split == BL, might be error (no actual split found)
            if abs(split_float - bl_float) < 0.001:
                logging.warning(f"‚ö†Ô∏è VLM extracted split qty equals B/L: {split_value} == {bl_value} ‚Üí returning None")
                return None

            # Valid split quantity
            logging.info(f"‚úì VLM extracted valid split qty: {split_value} <= {bl_value}")
            return split_value

        except Exception as e:
            logging.warning(f"Error validating split qty: {e}")
            return None

    if provider == 'local':
        try:
            from analyzers.clients.ollama_client import call_ollama_vlm
            text, tokens = call_ollama_vlm(file_path, field_name, lang, model=ollama_model)

            # Clean VOY.NO. if needed
            if field_name == 'VOY.NO.':
                original = text
                text = clean_voy_no(text)
                if original != text:
                    logging.info(f"üßπ VLM Cleaned VOY.NO.: '{original}' ‚Üí '{text}'")

            # Validate split quantity if needed
            elif field_name == 'B/L split quantity (MT)' and bl_quantity:
                original = text
                text = validate_split_qty(text, bl_quantity)
                if original != text:
                    logging.info(f"üßπ VLM Validated split qty: '{original}' ‚Üí '{text}'")

            return text, tokens
        except Exception as e:
            logging.warning(f"Ollama VLM Â§±ÊïóÔºåÊîπÁî®Èõ≤Á´ØÔºö{e}")
            from analyzers.services import vlm_fill_field as _vlm
            text, tokens = _vlm(file_path, field_name, lang)

            # Clean VOY.NO. if needed
            if field_name == 'VOY.NO.':
                original = text
                text = clean_voy_no(text)
                if original != text:
                    logging.info(f"üßπ VLM Cleaned VOY.NO.: '{original}' ‚Üí '{text}'")

            # Validate split quantity if needed
            elif field_name == 'B/L split quantity (MT)' and bl_quantity:
                original = text
                text = validate_split_qty(text, bl_quantity)
                if original != text:
                    logging.info(f"üßπ VLM Validated split qty: '{original}' ‚Üí '{text}'")

            return text, tokens
    else:
        from analyzers.services import vlm_fill_field as _vlm
        text, tokens = _vlm(file_path, field_name, lang)

        # Clean VOY.NO. if needed
        if field_name == 'VOY.NO.':
            original = text
            text = clean_voy_no(text)
            if original != text:
                logging.info(f"üßπ VLM Cleaned VOY.NO.: '{original}' ‚Üí '{text}'")

        # Validate split quantity if needed
        elif field_name == 'B/L split quantity (MT)' and bl_quantity:
            original = text
            text = validate_split_qty(text, bl_quantity)
            if original != text:
                logging.info(f"üßπ VLM Validated split qty: '{original}' ‚Üí '{text}'")

        return text, tokens

# =========================
# ÁÆ°ÁêÜËÄÖÂæåÂè∞Ôºö‰ΩøÁî®ËÄÖËàáÂàÜÊûêÂô®Ë®≠ÂÆö„ÄÅÊ≠∑Âè≤
# =========================

@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    rows = conn.execute('SELECT id, username, display_name, is_admin, is_active, email, phone, address, logo_file, notes FROM users ORDER BY id ASC').fetchall()
    conn.close()
    users = [dict(r) for r in rows]
    
    # ÈáçÊñ∞Âä†ËºâÊúÄÊñ∞ÁöÑÊ≠∑Âè≤Ë®òÈåÑ
    global history
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'rb') as f:
                history = pickle.load(f)
        except Exception as e:
            app.logger.error(f"Failed to reload history: {e}")
    
    # ÂÇ≥ÈÅûÊ≠∑Âè≤Ë≥áÊñô‰æõÁµ±Ë®à‰ΩøÁî®ÔºåÊåâÊôÇÈñìÈôçÂ∫èÊéíÂ∫è
    def get_history_time(x):
        return x.get('time') or x.get('timestamp') or ''
    sorted_history = sorted(history, key=get_history_time, reverse=True)
    
    # Create user mapping table (username -> user info)
    user_map = {u['username']: u for u in users}
    
    # Calculate statistics
    stats = calculate_user_statistics(sorted_history)
    
    return render_template('admin_dashboard.html', users=users, history=sorted_history, user_map=user_map, selected_username='', stats=stats)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    username = request.form.get('username', '').strip()
    display_name = request.form.get('display_name', '').strip()
    password = request.form.get('password', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    notes = request.form.get('notes', '').strip()
    is_admin = 1 if request.form.get('is_admin') == 'on' else 0
    
    if not username or not password:
        return redirect(url_for('admin_dashboard'))
    
    # ËôïÁêÜ logo Êñá‰ª∂‰∏äÂÇ≥
    logo_filename = None
    if 'logo' in request.files:
        logo_file = request.files['logo']
        if logo_file and logo_file.filename:
            # Á¢∫‰øù uploads ÁõÆÈåÑÂ≠òÂú®
            import uuid
            import os
            from werkzeug.utils import secure_filename
            
            filename = secure_filename(logo_file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if file_ext in ['jpg', 'jpeg', 'png', 'gif']:
                logo_filename = f"user_logo_{uuid.uuid4().hex[:8]}.{file_ext}"
                logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                logo_file.save(logo_path)
    
    try:
        conn = get_db_connection()
        conn.execute(
            '''INSERT INTO users (username, display_name, password_hash, is_admin, is_active, email, phone, address, logo_file, notes) 
               VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?, ?)''',
            (username, display_name or None, generate_password_hash(password, method='pbkdf2:sha256'), is_admin, 
             email or None, phone or None, address or None, logo_filename, notes or None)
        )
        conn.commit()
    except Exception as e:
        logging.error(f"Failed to add user: {e}")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/edit', methods=['POST'])
@admin_required
def admin_edit_user():
    user_id = request.form.get('user_id')
    username = request.form.get('username', '').strip()
    display_name = request.form.get('display_name', '').strip()
    password = request.form.get('password', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    notes = request.form.get('notes', '').strip()
    is_admin = 1 if request.form.get('is_admin') == 'on' else 0
    
    if not user_id or not username:
        return redirect(url_for('admin_dashboard'))
    
    # ËôïÁêÜ logo Êñá‰ª∂‰∏äÂÇ≥
    logo_filename = None
    if 'logo' in request.files:
        logo_file = request.files['logo']
        if logo_file and logo_file.filename:
            import uuid
            import os
            from werkzeug.utils import secure_filename
            
            filename = secure_filename(logo_file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if file_ext in ['jpg', 'jpeg', 'png', 'gif']:
                logo_filename = f"user_logo_{uuid.uuid4().hex[:8]}.{file_ext}"
                logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                logo_file.save(logo_path)
    
    try:
        conn = get_db_connection()
        
        # Â¶ÇÊûúÊúâÊñ∞ÂØÜÁ¢ºÔºåÂâáÊõ¥Êñ∞ÂØÜÁ¢º
        if password:
            if logo_filename:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, password_hash=?, is_admin=?, email=?, phone=?, address=?, logo_file=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, generate_password_hash(password, method='pbkdf2:sha256'), is_admin, 
                     email or None, phone or None, address or None, logo_filename, notes or None, user_id)
                )
            else:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, password_hash=?, is_admin=?, email=?, phone=?, address=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, generate_password_hash(password, method='pbkdf2:sha256'), is_admin, 
                     email or None, phone or None, address or None, notes or None, user_id)
                )
        else:
            # ‰∏çÊõ¥Êñ∞ÂØÜÁ¢º
            if logo_filename:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, is_admin=?, email=?, phone=?, address=?, logo_file=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, is_admin, email or None, phone or None, address or None, logo_filename, notes or None, user_id)
                )
            else:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, is_admin=?, email=?, phone=?, address=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, is_admin, email or None, phone or None, address or None, notes or None, user_id)
                )
        
        conn.commit()
    except Exception as e:
        logging.error(f"Á∑®ËºØ‰ΩøÁî®ËÄÖÂ§±Êïó: {e}")
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/<int:user_id>/toggle_active', methods=['POST'])
@admin_required
def admin_toggle_user_active(user_id: int):
    conn = get_db_connection()
    row = conn.execute('SELECT username, is_active FROM users WHERE id=?', (user_id,)).fetchone()
    if row:
        if row['username'] == 'admin':
            conn.close()
            return redirect(url_for('admin_dashboard'))
        new_active = 0 if row['is_active'] == 1 else 1
        conn.execute('UPDATE users SET is_active=? WHERE id=?', (new_active, user_id))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id: int):
    conn = get_db_connection()
    row = conn.execute('SELECT username FROM users WHERE id=?', (user_id,)).fetchone()
    if row and row['username'] != 'admin':
        conn.execute('DELETE FROM user_analyzers WHERE user_id=?', (user_id,))
        conn.execute('DELETE FROM users WHERE id=?', (user_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

def _get_user_analyzer_settings_map(user_id: int) -> Dict[str, Dict[str, Any]]:
    conn = get_db_connection()
    rows = conn.execute('SELECT analyzer, enabled, vlm_provider, ollama_model, ocr_lang, save_files FROM user_analyzers WHERE user_id=?', (user_id,)).fetchall()
    conn.close()
    settings = {}
    for r in rows:
        settings[r['analyzer']] = {
            'enabled': bool(r['enabled']),
            'vlm_provider': r['vlm_provider'],
            'ollama_model': r['ollama_model'],
            'ocr_lang': r['ocr_lang'] or 'auto',  # ÈªòË™çËá™ÂãïÂÅµÊ∏¨
            'save_files': bool(r['save_files'])
        }
    return settings

@app.route('/admin/users/<int:user_id>/analyzers', methods=['GET', 'POST'])
@admin_required
def admin_user_analyzers(user_id: int):
    conn = get_db_connection()
    user = conn.execute('SELECT id, username FROM users WHERE id=?', (user_id,)).fetchone()
    conn.close()
    if not user:
        return redirect(url_for('admin_dashboard'))
    analyzers_list = list(config.get('prompts', {}).keys())
    if request.method == 'POST':
        conn = get_db_connection()
        for analyzer in analyzers_list:
            enabled = 1 if request.form.get(f'enabled_{analyzer}') == 'on' else 0
            provider = request.form.get(f'provider_{analyzer}', 'cloud')
            model = request.form.get(f'model_{analyzer}', None)
            ocr_lang = request.form.get(f'ocr_lang_{analyzer}', 'auto')
            save_files = 1 if request.form.get(f'save_files_{analyzer}') == 'on' else 0
            exist = conn.execute('SELECT id FROM user_analyzers WHERE user_id=? AND analyzer=?', (user_id, analyzer)).fetchone()
            if exist:
                conn.execute(
                    'UPDATE user_analyzers SET enabled=?, vlm_provider=?, ollama_model=?, ocr_lang=?, save_files=? WHERE user_id=? AND analyzer=?',
                    (enabled, provider, model, ocr_lang, save_files, user_id, analyzer)
                )
            else:
                conn.execute(
                    'INSERT INTO user_analyzers (user_id, analyzer, enabled, vlm_provider, ollama_model, ocr_lang, save_files) VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (user_id, analyzer, enabled, provider, model, ocr_lang, save_files)
                )
        conn.commit()
        conn.close()
        return redirect(url_for('admin_user_analyzers', user_id=user_id) + '?saved=1')
    # GET
    settings = _get_user_analyzer_settings_map(user_id)
    return render_template('admin_user_analyzers.html', user=dict(user), analyzers=analyzers_list, settings=settings)

@app.route('/admin/history')
@admin_required
def admin_history():
    selected_username = request.args.get('username', '').strip()
    items = history
    if selected_username:
        items = [h for h in history if h.get('username') == selected_username]
    # ÊéíÂ∫è
    def get_history_time(x):
        return x.get('time') or x.get('timestamp') or ''
    items_sorted = sorted(items, key=get_history_time, reverse=True)
    # ÈúÄË¶Å users Ê∏ÖÂñÆ‰æõ‰∏ãÊãâ
    conn = get_db_connection()
    rows = conn.execute('SELECT id, username, display_name, is_admin, is_active, email, phone, address, logo_file, notes FROM users ORDER BY id ASC').fetchall()
    conn.close()
    users = [dict(r) for r in rows]
    
    # Create user mapping table (username -> user info)
    user_map = {u['username']: u for u in users}
    
    # Calculate statistics (for filtered records)
    stats = calculate_user_statistics(items_sorted)
    
    return render_template('admin_dashboard.html', users=users, history=items_sorted, user_map=user_map, selected_username=selected_username, stats=stats)

@app.route('/admin/history/delete', methods=['POST'])
@admin_required
def admin_delete_history():
    try:
        data = request.get_json()
        username = data.get('username')
        timestamp = data.get('timestamp')
        excel_file = data.get('excel_file')
        
        if not username or not timestamp:
            return jsonify({'success': False, 'error': 'Missing required parameters'})
        
        # Find and delete corresponding records from history
        global history
        original_count = len(history)
        
        # Find records to delete
        to_remove = []
        for i, h in enumerate(history):
            if (h.get('username') == username and 
                h.get('time') == timestamp):
                to_remove.append(i)
                
                # Âà™Èô§Áõ∏ÈóúÊ™îÊ°à
                if excel_file and h.get('excel') == excel_file:
                    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file)
                    try:
                        if os.path.exists(excel_path):
                            os.remove(excel_path)
                            app.logger.info(f"Âà™Èô§ Excel Ê™îÊ°à: {excel_path}")
                    except Exception as e:
                        app.logger.warning(f"Âà™Èô§ Excel Ê™îÊ°àÂ§±Êïó: {e}")
                
                # Âà™Èô§Êó•Ë™åÊ™îÊ°à
                log_file = h.get('log_file')
                if log_file:
                    log_path = os.path.join('logs', log_file)
                    try:
                        if os.path.exists(log_path):
                            os.remove(log_path)
                            app.logger.info(f"Âà™Èô§Êó•Ë™åÊ™îÊ°à: {log_path}")
                    except Exception as e:
                        app.logger.warning(f"Âà™Èô§Êó•Ë™åÊ™îÊ°àÂ§±Êïó: {e}")
        
        # Remove from history records (reverse order to avoid index issues)
        for i in sorted(to_remove, reverse=True):
            del history[i]
        
        # Save updated history records
        save_history()
        
        deleted_count = len(to_remove)
        if deleted_count > 0:
            app.logger.info(f"Administrator deleted {deleted_count} history records for user {username}")
            return jsonify({'success': True, 'message': f'Successfully deleted {deleted_count} records'})
        else:
            return jsonify({'success': False, 'error': 'No matching records found'})
            
    except Exception as e:
        app.logger.error(f"Failed to delete history records: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/history/clear-user', methods=['POST'])
@admin_required
def admin_clear_user_history():
    try:
        data = request.get_json()
        username = data.get('username')
        
        if not username:
            return jsonify({'success': False, 'error': 'Missing username parameter'})
        
        global history
        original_count = len(history)
        
        # Find records to delete
        to_remove = []
        for i, h in enumerate(history):
            if h.get('username') == username:
                to_remove.append(i)
                
                # Delete related files
                excel_file = h.get('excel')
                if excel_file:
                    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file)
                    try:
                        if os.path.exists(excel_path):
                            os.remove(excel_path)
                            app.logger.info(f"Deleted Excel file: {excel_path}")
                    except Exception as e:
                        app.logger.warning(f"Failed to delete Excel file: {e}")
                
                # Delete log file
                log_file = h.get('log_file')
                if log_file:
                    log_path = os.path.join('logs', log_file)
                    try:
                        if os.path.exists(log_path):
                            os.remove(log_path)
                            app.logger.info(f"Deleted log file: {log_path}")
                    except Exception as e:
                        app.logger.warning(f"Failed to delete log file: {e}")
        
        # Remove from history records (reverse order to avoid index issues)
        for i in sorted(to_remove, reverse=True):
            del history[i]
        
        # Save updated history records
        save_history()
        
        deleted_count = len(to_remove)
        app.logger.info(f"Administrator cleared {deleted_count} history records for user {username}")
        return jsonify({'success': True, 'deleted_count': deleted_count})
            
    except Exception as e:
        app.logger.error(f"Failed to clear user history records: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/history/clear-all', methods=['POST'])
@admin_required
def admin_clear_all_history():
    try:
        global history
        original_count = len(history)
        
        # Delete all related files
        for h in history:
            # Delete Excel file
            excel_file = h.get('excel')
            if excel_file:
                excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file)
                try:
                    if os.path.exists(excel_path):
                        os.remove(excel_path)
                        app.logger.info(f"Deleted Excel file: {excel_path}")
                except Exception as e:
                    app.logger.warning(f"Failed to delete Excel file: {e}")
            
            # Delete log file
            log_file = h.get('log_file')
            if log_file:
                log_path = os.path.join('logs', log_file)
                try:
                    if os.path.exists(log_path):
                        os.remove(log_path)
                        app.logger.info(f"Deleted log file: {log_path}")
                except Exception as e:
                    app.logger.warning(f"Failed to delete log file: {e}")
        
        # Clear all history records
        history.clear()
        save_history()
        
        app.logger.info(f"Administrator cleared all {original_count} history records")
        return jsonify({'success': True, 'deleted_count': original_count})
            
    except Exception as e:
        app.logger.error(f"Failed to clear all history records: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/ollama/models')
@admin_required
def admin_ollama_models():
    try:
        from analyzers.clients.ollama_client import get_ollama_models
        models = get_ollama_models()
        return jsonify({'success': True, 'models': models})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'models': []}), 500

@app.route('/admin/user-stats/<username>')
@admin_required
def admin_user_stats(username):
    """Get statistics for a specific user"""
    try:
        # Get the user's history records
        user_history = [h for h in history if h.get('username') == username]
        
        # Calculate statistics
        stats = calculate_user_statistics(user_history)
        
        if username in stats:
            return jsonify({
                'success': True,
                'stats': stats[username]
            })
        else:
            return jsonify({
                'success': True,
                'stats': {
                    'total_tokens': 0,
                    'total_records': 0,
                    'daily_usage': {}
                }
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
# ÁßªÈô§Ë™ûË®ÄÂàáÊèõË∑ØÁî±

if __name__ == '__main__':
    init_db()

    # Get configuration from environment variables
    debug_mode = os.environ.get('FLASK_ENV', 'development') == 'development'
    port = int(os.environ.get('PORT', 5001))

    # Always bind to 0.0.0.0 in container so it's accessible externally
    host = '0.0.0.0'

    app.run(debug=debug_mode, host=host, port=port)
